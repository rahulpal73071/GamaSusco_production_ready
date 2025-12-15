"""
Reports & Analytics Router
Generate comprehensive emission reports with PDF and Excel exports
"""
from fastapi import APIRouter, Depends, HTTPException, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, extract
from typing import List, Optional, Dict, Any
from datetime import datetime, date, timedelta
from pydantic import BaseModel
import pandas as pd
import io
import json
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# PDF Generation imports
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Chart generation imports
import matplotlib

matplotlib.use('Agg')
import matplotlib.pyplot as plt

from app.database import get_db
from app.models import EmissionActivity, Company, User, Goal, AIRecommendation
from app.routers.auth import get_current_user
from app.db_maintenance import ensure_ai_recommendation_schema

router = APIRouter(prefix="/api/v1/reports", tags=["Reports & Analytics"])


# ==================== PYDANTIC MODELS ====================

class CompanyProfile(BaseModel):
    company_name: str
    address: Optional[str] = "Address not set"
    industry: Optional[str] = "Industry not set"
    reporting_officer: str
    email: str
    phone: Optional[str] = "N/A"
    website: Optional[str] = None
    logo_url: Optional[str] = None


class DateRange(BaseModel):
    start_date: date
    end_date: date


class ReportConfig(BaseModel):
    report_type: str  # summary, detailed, comparison, trend, comprehensive
    date_range: DateRange
    scope: Optional[int] = None
    category: Optional[str] = None
    include_water: bool = True
    include_waste: bool = True
    include_scope3: bool = True
    group_by: str = "month"  # day, week, month, quarter, year
    export_format: str = "json"  # json, csv, excel, pdf


class SummaryMetrics(BaseModel):
    total_emissions_kg: float
    total_emissions_tons: float
    scope1_emissions: float
    scope2_emissions: float
    scope3_emissions: float
    top_category: str
    top_category_emissions: float
    average_daily_emissions: float
    total_activities: int
    comparison_to_previous_period: float


class TrendData(BaseModel):
    period: str
    emissions_kg: float
    activities_count: int
    average_emission_per_activity: float


class CategoryBreakdown(BaseModel):
    category: str
    emissions_kg: float
    emissions_percent: float
    activities_count: int


class DetailedReport(BaseModel):
    summary: SummaryMetrics
    trends: List[TrendData]
    category_breakdown: List[CategoryBreakdown]
    scope_breakdown: Dict[str, float]
    top_emission_sources: List[Dict[str, Any]]
    goals_progress: List[Dict[str, Any]]
    recommendations_summary: Dict[str, Any]


class ComprehensiveReportData(BaseModel):
    company_profile: CompanyProfile
    summary_metrics: Dict[str, Any]
    emissions_data: Dict[str, Any]
    water_data: Dict[str, Any]
    waste_data: Dict[str, Any]
    scope3_data: Dict[str, Any]
    trends: List[Dict[str, Any]]
    recommendations: List[Dict[str, Any]]  # Changed to dict for AI recommendations
    goals: List[Dict[str, Any]] = []  # Added goals
    activities: List[Dict[str, Any]] = []  # Added top activities
    category_breakdown: List[Dict[str, Any]] = []  # Added category breakdown
    compliance_info: Dict[str, Any] = {}  # Added compliance information
    generated_at: datetime


# ==================== HELPER FUNCTIONS ====================

def calculate_summary_metrics(
        db: Session,
        company_id: int,
        start_date: date,
        end_date: date,
        scope: Optional[int] = None,
        category: Optional[str] = None
) -> SummaryMetrics:
    """Calculate summary metrics for a date range"""

    # Base query
    query = db.query(EmissionActivity).filter(
        EmissionActivity.company_id == company_id,
        EmissionActivity.activity_date >= start_date,
        EmissionActivity.activity_date <= end_date
    )

    if scope:
        query = query.filter(EmissionActivity.scope_number == scope)
    if category:
        query = query.filter(EmissionActivity.category.ilike(f"%{category}%"))

    records = query.all()

    # Calculate totals
    total_emissions = sum(r.emissions_kgco2e for r in records)
    scope1 = sum(r.emissions_kgco2e for r in records if r.scope_number == 1)
    scope2 = sum(r.emissions_kgco2e for r in records if r.scope_number == 2)
    scope3 = sum(r.emissions_kgco2e for r in records if r.scope_number == 3)

    # Category breakdown
    category_totals = {}
    for r in records:
        cat = r.category or "Uncategorized"
        category_totals[cat] = category_totals.get(cat, 0) + r.emissions_kgco2e

    top_category = max(category_totals.items(), key=lambda x: x[1]) if category_totals else ("N/A", 0)

    # Calculate average daily emissions
    days_diff = (end_date - start_date).days + 1
    avg_daily = total_emissions / days_diff if days_diff > 0 else 0

    # Compare to previous period
    period_length = (end_date - start_date).days + 1
    previous_start = start_date - timedelta(days=period_length)
    previous_end = start_date - timedelta(days=1)

    previous_query = db.query(func.sum(EmissionActivity.emissions_kgco2e)).filter(
        EmissionActivity.company_id == company_id,
        EmissionActivity.activity_date >= previous_start,
        EmissionActivity.activity_date <= previous_end
    )

    if scope:
        previous_query = previous_query.filter(EmissionActivity.scope_number == scope)
    if category:
        previous_query = previous_query.filter(EmissionActivity.category.ilike(f"%{category}%"))

    previous_emissions = previous_query.scalar() or 0

    comparison = ((total_emissions - previous_emissions) / previous_emissions * 100) if previous_emissions > 0 else 0

    return SummaryMetrics(
        total_emissions_kg=round(total_emissions, 2),
        total_emissions_tons=round(total_emissions / 1000, 2),
        scope1_emissions=round(scope1, 2),
        scope2_emissions=round(scope2, 2),
        scope3_emissions=round(scope3, 2),
        top_category=top_category[0],
        top_category_emissions=round(top_category[1], 2),
        average_daily_emissions=round(avg_daily, 2),
        total_activities=len(records),
        comparison_to_previous_period=round(comparison, 2)
    )


def generate_trend_data(
        db: Session,
        company_id: int,
        start_date: date,
        end_date: date,
        group_by: str,
        scope: Optional[int] = None,
        category: Optional[str] = None
) -> List[TrendData]:
    """Generate trend data grouped by specified period"""

    query = db.query(EmissionActivity).filter(
        EmissionActivity.company_id == company_id,
        EmissionActivity.activity_date >= start_date,
        EmissionActivity.activity_date <= end_date
    )

    if scope:
        query = query.filter(EmissionActivity.scope_number == scope)
    if category:
        query = query.filter(EmissionActivity.category.ilike(f"%{category}%"))

    records = query.all()

    # Group data
    grouped_data = {}

    for record in records:
        if not record.activity_date:
            continue

        if group_by == "day":
            period_key = record.activity_date.strftime("%Y-%m-%d")
        elif group_by == "week":
            period_key = record.activity_date.strftime("%Y-W%U")
        elif group_by == "month":
            period_key = record.activity_date.strftime("%Y-%m")
        elif group_by == "quarter":
            quarter = (record.activity_date.month - 1) // 3 + 1
            period_key = f"{record.activity_date.year}-Q{quarter}"
        elif group_by == "year":
            period_key = str(record.activity_date.year)
        else:
            period_key = record.activity_date.strftime("%Y-%m")

        if period_key not in grouped_data:
            grouped_data[period_key] = {
                "emissions": 0,
                "count": 0
            }

        grouped_data[period_key]["emissions"] += record.emissions_kgco2e
        grouped_data[period_key]["count"] += 1

    # Convert to list
    trends = []
    for period, data in sorted(grouped_data.items()):
        avg_emission = data["emissions"] / data["count"] if data["count"] > 0 else 0
        trends.append(TrendData(
            period=period,
            emissions_kg=round(data["emissions"], 2),
            activities_count=data["count"],
            average_emission_per_activity=round(avg_emission, 2)
        ))

    return trends


def generate_category_breakdown(
        db: Session,
        company_id: int,
        start_date: date,
        end_date: date,
        scope: Optional[int] = None
) -> List[CategoryBreakdown]:
    """Generate category breakdown"""

    query = db.query(
        EmissionActivity.category,
        func.sum(EmissionActivity.emissions_kgco2e).label('total_emissions'),
        func.count(EmissionActivity.id).label('count')
    ).filter(
        EmissionActivity.company_id == company_id,
        EmissionActivity.activity_date >= start_date,
        EmissionActivity.activity_date <= end_date
    )

    if scope:
        query = query.filter(EmissionActivity.scope_number == scope)

    results = query.group_by(EmissionActivity.category).all()

    total_emissions = sum(r.total_emissions for r in results)

    breakdown = []
    for category, emissions, count in results:
        cat = category or "Uncategorized"
        percent = (emissions / total_emissions * 100) if total_emissions > 0 else 0
        breakdown.append(CategoryBreakdown(
            category=cat,
            emissions_kg=round(emissions, 2),
            emissions_percent=round(percent, 2),
            activities_count=count
        ))

    # Sort by emissions descending
    breakdown.sort(key=lambda x: x.emissions_kg, reverse=True)

    return breakdown


def get_water_metrics(db: Session, company_id: int, start_date: date, end_date: date) -> Dict[str, Any]:
    """Get water usage metrics from actual WaterUsage data"""
    from app.models import WaterUsage
    from sqlalchemy import func
    
    try:
        # Query water usage records within date range
        water_records = db.query(WaterUsage).filter(
            WaterUsage.company_id == company_id,
            WaterUsage.created_at >= datetime.combine(start_date, datetime.min.time()),
            WaterUsage.created_at <= datetime.combine(end_date, datetime.max.time())
        ).all()
        
        print(f"🔍 Water metrics query: company_id={company_id}, date_range={start_date} to {end_date}, found {len(water_records)} records")
        
        if not water_records or len(water_records) == 0:
            # Return zeros if no data
            print(f"✅ No water records found for company {company_id}, returning zeros")
            return {
                "total_usage": 0.0,
                "municipal_water": 0.0,
                "groundwater": 0.0,
                "rainwater": 0.0,
                "surface_water": 0.0,
                "avg_daily": 0.0,
                "compliance_status": "No Data Available"
            }
        
        # Calculate totals by source
        total_usage = 0.0
        municipal_water = 0.0
        groundwater = 0.0
        rainwater = 0.0
        surface_water = 0.0
        
        for record in water_records:
            withdrawal = record.withdrawal_volume or 0.0
            total_usage += withdrawal
            
            # Group by source
            source = (record.source or "").lower()
            if "municipal" in source:
                municipal_water += withdrawal
            elif "ground" in source:
                groundwater += withdrawal
            elif "rain" in source:
                rainwater += withdrawal
            elif "surface" in source:
                surface_water += withdrawal
            else:
                # Default to municipal if source is unclear
                municipal_water += withdrawal
        
        # If all values are zero (no actual data), return zeros
        if total_usage == 0.0:
            print(f"✅ Water records found but all values are zero, returning zeros")
            return {
                "total_usage": 0.0,
                "municipal_water": 0.0,
                "groundwater": 0.0,
                "rainwater": 0.0,
                "surface_water": 0.0,
                "avg_daily": 0.0,
                "compliance_status": "No Data Available"
            }
        
        # Calculate average daily usage
        days = (end_date - start_date).days + 1
        avg_daily = total_usage / days if days > 0 else 0.0
        
        return {
            "total_usage": round(total_usage, 2),
            "municipal_water": round(municipal_water, 2),
            "groundwater": round(groundwater, 2),
            "rainwater": round(rainwater, 2),
            "surface_water": round(surface_water, 2),
            "avg_daily": round(avg_daily, 2),
            "compliance_status": "Compliant" if total_usage > 0 else "No Data Available"
        }
    except Exception as e:
        print(f"⚠️ Error fetching water metrics: {e}")
        # Return zeros on error
        return {
            "total_usage": 0.0,
            "municipal_water": 0.0,
            "groundwater": 0.0,
            "rainwater": 0.0,
            "surface_water": 0.0,
            "avg_daily": 0.0,
            "compliance_status": "Error Fetching Data"
    }


def get_waste_metrics(db: Session, company_id: int, start_date: date, end_date: date) -> Dict[str, Any]:
    """Get waste management metrics from actual WasteDisposal data"""
    from app.models import WasteDisposal
    from sqlalchemy import func
    
    try:
        # Query waste disposal records within date range
        waste_records = db.query(WasteDisposal).filter(
            WasteDisposal.company_id == company_id,
            WasteDisposal.created_at >= datetime.combine(start_date, datetime.min.time()),
            WasteDisposal.created_at <= datetime.combine(end_date, datetime.max.time())
        ).all()
        
        print(f"🔍 Waste metrics query: company_id={company_id}, date_range={start_date} to {end_date}, found {len(waste_records)} records")
        
        if not waste_records or len(waste_records) == 0:
            # Return zeros if no data
            print(f"✅ No waste records found for company {company_id}, returning zeros")
            return {
                "total_waste": 0.0,
                "recycled": 0.0,
                "landfill": 0.0,
                "incinerated": 0.0,
                "composted": 0.0,
                "recycling_rate": 0.0,
                "compliance_status": "No Data Available"
            }
        
        # Calculate totals by disposal method
        total_waste = 0.0
        recycled = 0.0
        landfill = 0.0
        incinerated = 0.0
        composted = 0.0
        
        for record in waste_records:
            quantity = record.quantity or 0.0
            total_waste += quantity
            
            # Group by disposal method
            method = (record.disposal_method or "").lower()
            if "recycl" in method:
                recycled += quantity
            elif "landfill" in method or "dump" in method:
                landfill += quantity
            elif "incinerat" in method or "burn" in method:
                incinerated += quantity
            elif "compost" in method:
                composted += quantity
            else:
                # Default to landfill if method is unclear
                landfill += quantity
        
        # If all values are zero (no actual data), return zeros
        if total_waste == 0.0:
            print(f"✅ Waste records found but all values are zero, returning zeros")
            return {
                "total_waste": 0.0,
                "recycled": 0.0,
                "landfill": 0.0,
                "incinerated": 0.0,
                "composted": 0.0,
                "recycling_rate": 0.0,
                "compliance_status": "No Data Available"
            }
        
        # Calculate recycling rate
        recycling_rate = (recycled / total_waste * 100) if total_waste > 0 else 0.0
        
        return {
            "total_waste": round(total_waste, 2),
            "recycled": round(recycled, 2),
            "landfill": round(landfill, 2),
            "incinerated": round(incinerated, 2),
            "composted": round(composted, 2),
            "recycling_rate": round(recycling_rate, 2),
            "compliance_status": "Compliant" if total_waste > 0 else "No Data Available"
        }
    except Exception as e:
        print(f"⚠️ Error fetching waste metrics: {e}")
        # Return zeros on error
        return {
            "total_waste": 0.0,
            "recycled": 0.0,
            "landfill": 0.0,
            "incinerated": 0.0,
            "composted": 0.0,
            "recycling_rate": 0.0,
            "compliance_status": "Error Fetching Data"
    }


def get_scope3_metrics(db: Session, company_id: int, start_date: date, end_date: date) -> Dict[str, Any]:
    """Get Scope 3 breakdown metrics"""

    query = db.query(EmissionActivity).filter(
        EmissionActivity.company_id == company_id,
        EmissionActivity.scope_number == 3,
        EmissionActivity.activity_date >= start_date,
        EmissionActivity.activity_date <= end_date
    )

    records = query.all()

    # Group by category
    breakdown = {}
    for r in records:
        cat = r.category or "Other"
        breakdown[cat] = breakdown.get(cat, 0) + r.emissions_kgco2e

    return breakdown


def generate_emissions_chart(data: Dict[str, Any]) -> io.BytesIO:
    """Generate emissions breakdown pie chart"""

    fig, ax = plt.subplots(figsize=(6, 4))

    scopes = ['Scope 1', 'Scope 2', 'Scope 3']
    values = [
        data.get('scope1_emissions', 0),
        data.get('scope2_emissions', 0),
        data.get('scope3_emissions', 0)
    ]
    colors_list = ['#ef4444', '#f59e0b', '#10b981']

    # Only plot if there's data
    if sum(values) > 0:
        ax.pie(values, labels=scopes, autopct='%1.1f%%', colors=colors_list, startangle=90)
        ax.set_title('Emissions by Scope', fontsize=14, fontweight='bold')
    else:
        ax.text(0.5, 0.5, 'No emission data', ha='center', va='center')

    # Save to bytes
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    buf.seek(0)
    plt.close()

    return buf


def generate_trend_chart(trends: List[Dict[str, Any]]) -> io.BytesIO:
    """Generate emissions trend line chart"""

    fig, ax = plt.subplots(figsize=(8, 4))

    if trends and len(trends) > 0:
        periods = [t['period'] for t in trends]
        emissions = [t['emissions_kg'] for t in trends]

        ax.plot(periods, emissions, marker='o', linewidth=2, color='#667eea')
        ax.set_xlabel('Period', fontsize=10)
        ax.set_ylabel('Emissions (kg CO2e)', fontsize=10)
        ax.set_title('Emissions Trend', fontsize=14, fontweight='bold')
        ax.grid(True, alpha=0.3)

        # Rotate x labels if many periods
        if len(periods) > 6:
            plt.xticks(rotation=45, ha='right')
    else:
        ax.text(0.5, 0.5, 'No trend data', ha='center', va='center', transform=ax.transAxes)

    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    buf.seek(0)
    plt.close()

    return buf


def generate_pdf_report(report_data: ComprehensiveReportData) -> io.BytesIO:
    """Generate comprehensive PDF report"""

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4, 
        topMargin=0.75 * inch, 
        bottomMargin=0.75 * inch,
        leftMargin=0.75 * inch,
        rightMargin=0.75 * inch
    )
    story = []
    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=28,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )

    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=18,
        textColor=colors.HexColor('#374151'),
        spaceAfter=12,
        spaceBefore=20,
        fontName='Helvetica-Bold'
    )

    # ==================== PAGE 1: COVER PAGE ====================

    story.append(Spacer(1, 1.5 * inch))
    story.append(Paragraph("COMPREHENSIVE SUSTAINABILITY REPORT", title_style))
    story.append(Spacer(1, 0.5 * inch))

    # Company info
    company = report_data.company_profile
    company_info = f"""
    <para align=center>
    <b><font size=16>{company.company_name}</font></b><br/><br/>
    <font size=11>{company.address}</font><br/>
    <font size=11>Industry: {company.industry}</font><br/><br/>
    <font size=11>Reporting Officer: {company.reporting_officer}</font><br/>
    <font size=11>Email: {company.email}</font><br/>
    <font size=11>Phone: {company.phone}</font>
    </para>
    """
    story.append(Paragraph(company_info, styles['Normal']))
    story.append(Spacer(1, 0.5 * inch))

    # Reporting period
    period_text = f"""
    <para align=center>
    <b><font size=12>Reporting Period:</font></b><br/>
    <font size=11>{report_data.summary_metrics['reporting_period_start']} to {report_data.summary_metrics['reporting_period_end']}</font><br/><br/>
    <b><font size=12>Generated:</font></b><br/>
    <font size=11>{report_data.generated_at.strftime('%Y-%m-%d %H:%M:%S')}</font>
    </para>
    """
    story.append(Paragraph(period_text, styles['Normal']))
    story.append(Spacer(1, 0.3 * inch))

    # Standards compliance badge
    standards_text = """
    <para align=center>
    <font size=10 color=#059669><b>✓ GHG Protocol Compliant</b></font><br/>
    <font size=9>Aligned with ISO 14064 and GRI Standards</font>
    </para>
    """
    story.append(Paragraph(standards_text, styles['Normal']))
    story.append(PageBreak())

    # ==================== PAGE 2: EXECUTIVE SUMMARY ====================

    story.append(Paragraph("EXECUTIVE SUMMARY", heading_style))
    story.append(Spacer(1, 0.2 * inch))

    summary = report_data.summary_metrics

    exec_summary = f"""
    This comprehensive sustainability report provides a detailed overview of environmental performance 
    for <b>{company.company_name}</b> during the reporting period from {summary.get('reporting_period_start')} 
    to {summary.get('reporting_period_end')}. The report covers greenhouse gas emissions (Scope 1, 2, and 3), 
    water usage, waste management, and compliance metrics aligned with international standards.
    <br/><br/>
    <b>Key Performance Indicators:</b><br/>
    • Total GHG Emissions: <b>{summary.get('total_emissions_tons', 0):.2f} tons CO2e</b><br/>
    • Scope 1 (Direct): {summary.get('scope1_emissions', 0) / 1000:.2f} tons CO2e<br/>
    • Scope 2 (Indirect - Energy): {summary.get('scope2_emissions', 0) / 1000:.2f} tons CO2e<br/>
    • Scope 3 (Value Chain): {summary.get('scope3_emissions', 0) / 1000:.2f} tons CO2e<br/>
    • Total Water Usage: <b>{summary.get('total_water_usage_m3', 0):,.2f} m³</b><br/>
    • Total Waste Generated: <b>{summary.get('total_waste_kg', 0):,.2f} kg</b><br/>
    • Waste Recycling Rate: <b>{summary.get('recycling_rate', 0):.1f}%</b><br/>
    • Total Activities Tracked: <b>{summary.get('total_activities', 0)}</b><br/>
    • Reporting Period: <b>{summary.get('reporting_period_days', 0)} days</b>
    """
    story.append(Paragraph(exec_summary, styles['Normal']))
    story.append(Spacer(1, 0.3 * inch))

    # Key metrics table
    metrics_data = [
        ['Metric', 'Value', 'Unit'],
        ['Total Emissions', f"{summary.get('total_emissions_tons', 0):,.2f}", 'tons CO2e'],
        ['Total Water Usage', f"{summary.get('total_water_usage_m3', 0):,.2f}", 'm³'],
        ['Total Waste', f"{summary.get('total_waste_kg', 0):,.2f}", 'kg'],
        ['Recycling Rate', f"{summary.get('recycling_rate', 0):.1f}", '%'],
        ['Activities Tracked', f"{summary.get('total_activities', 0)}", 'count']
    ]

    metrics_table = Table(metrics_data, colWidths=[3.5 * inch, 2 * inch, 1.5 * inch])
    metrics_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')])
    ]))
    story.append(metrics_table)
    story.append(PageBreak())

    # ==================== PAGE 3: EMISSIONS BREAKDOWN ====================

    story.append(Paragraph("GREENHOUSE GAS EMISSIONS ANALYSIS", heading_style))
    story.append(Spacer(1, 0.2 * inch))

    # Emissions table
    total_emissions_kg = summary.get('total_emissions_kg', 0)
    emissions_data = [
        ['Scope', 'Emissions (kg CO2e)', 'Emissions (tons CO2e)', 'Percentage'],
        ['Scope 1 (Direct Emissions)', f"{summary.get('scope1_emissions', 0):,.2f}",
         f"{summary.get('scope1_emissions', 0) / 1000:.2f}",
         f"{(summary.get('scope1_emissions', 0) / total_emissions_kg * 100 if total_emissions_kg > 0 else 0):.1f}%"],
        ['Scope 2 (Indirect - Energy)', f"{summary.get('scope2_emissions', 0):,.2f}",
         f"{summary.get('scope2_emissions', 0) / 1000:.2f}",
         f"{(summary.get('scope2_emissions', 0) / total_emissions_kg * 100 if total_emissions_kg > 0 else 0):.1f}%"],
        ['Scope 3 (Value Chain)', f"{summary.get('scope3_emissions', 0):,.2f}",
         f"{summary.get('scope3_emissions', 0) / 1000:.2f}",
         f"{(summary.get('scope3_emissions', 0) / total_emissions_kg * 100 if total_emissions_kg > 0 else 0):.1f}%"],
        ['TOTAL', f"<b>{total_emissions_kg:,.2f}</b>", f"<b>{summary.get('total_emissions_tons', 0):.2f}</b>", '100.0%']
    ]

    emissions_table = Table(emissions_data, colWidths=[2.5 * inch, 1.8 * inch, 1.8 * inch, 1.2 * inch])
    emissions_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f3f4f6')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f9fafb')])
    ]))

    story.append(emissions_table)
    story.append(Spacer(1, 0.3 * inch))

    # Add enhanced emissions chart with context
    story.append(Spacer(1, 0.1 * inch))
    context_text = """
    <b>Understanding Scope Emissions:</b> The GHG Protocol categorizes emissions into three scopes. 
    Scope 1 includes direct emissions from sources owned or controlled by the company (e.g., fuel combustion, company vehicles). 
    Scope 2 covers indirect emissions from purchased electricity, heat, or steam. 
    Scope 3 includes all other indirect emissions that occur in the value chain (e.g., business travel, purchased goods, waste disposal).
    """
    story.append(Paragraph(context_text, styles['Normal']))
    story.append(Spacer(1, 0.15 * inch))
    try:
        from app.services.pdf_generator import PDFReportGenerator
        pdf_gen = PDFReportGenerator()
        emissions_dict = {
            'scope1': summary.get('scope1_emissions', 0),
            'scope2': summary.get('scope2_emissions', 0),
            'scope3': summary.get('scope3_emissions', 0)
        }
        chart_img = pdf_gen._create_scope_pie_chart(emissions_dict)
        if chart_img:
            story.append(chart_img)
    except Exception as e:
        story.append(Paragraph(f"<i>Chart generation error: {str(e)}</i>", styles['Normal']))

    story.append(Spacer(1, 0.2 * inch))

    # Category breakdown
    story.append(Paragraph("Emissions by Category", heading_style))
    story.append(Spacer(1, 0.1 * inch))
    if report_data.category_breakdown and len(report_data.category_breakdown) > 0:
        category_data = [['Category', 'Emissions (tons CO2e)', 'Percentage', 'Activities']]
        for cat in report_data.category_breakdown[:8]:  # Top 8 categories
            category_data.append([
                cat.get('category', 'Uncategorized'),
                f"{cat.get('emissions_kg', 0) / 1000:.2f}",
                f"{cat.get('emissions_percent', 0):.1f}%",
                str(cat.get('activities_count', 0))
            ])

        category_table = Table(category_data, colWidths=[2.5 * inch, 1.5 * inch, 1.2 * inch, 1.2 * inch])
        category_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0fdf4')])
        ]))
        story.append(category_table)
    else:
        story.append(Paragraph("<i>No category breakdown available. No activities recorded for this reporting period.</i>", styles['Normal']))

    story.append(PageBreak())

    # ==================== PAGE 4: WATER & WASTE MANAGEMENT ====================

    story.append(Paragraph("WATER MANAGEMENT", heading_style))
    story.append(Spacer(1, 0.1 * inch))

    water_data = report_data.water_data

    water_table_data = [
        ['Metric', 'Value', 'Unit'],
        ['Total Water Usage', f"{water_data.get('total_usage', 0):,.2f}", 'm³'],
        ['Municipal Water', f"{water_data.get('municipal_water', 0):,.2f}", 'm³'],
        ['Groundwater', f"{water_data.get('groundwater', 0):,.2f}", 'm³'],
        ['Average Daily Usage', f"{water_data.get('avg_daily', 0):.2f}", 'm³/day'],
        ['Compliance Status', water_data.get('compliance_status', 'N/A'), '-']
    ]

    water_table = Table(water_table_data, colWidths=[3 * inch, 2.5 * inch, 1.5 * inch])
    water_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#eff6ff')])
    ]))

    story.append(water_table)
    story.append(Spacer(1, 0.4 * inch))

    story.append(Paragraph("WASTE MANAGEMENT", heading_style))
    story.append(Spacer(1, 0.1 * inch))

    waste_data = report_data.waste_data

    waste_table_data = [
        ['Metric', 'Value', 'Unit'],
        ['Total Waste Generated', f"{waste_data.get('total_waste', 0):,.2f}", 'kg'],
        ['Recycled Waste', f"{waste_data.get('recycled', 0):,.2f}", 'kg'],
        ['Landfill Waste', f"{waste_data.get('landfill', 0):,.2f}", 'kg'],
        ['Incinerated Waste', f"{waste_data.get('incinerated', 0):,.2f}", 'kg'],
        ['Recycling Rate', f"{waste_data.get('recycling_rate', 0):.1f}", '%'],
        ['Compliance Status', waste_data.get('compliance_status', 'N/A'), '-']
    ]

    waste_table = Table(waste_table_data, colWidths=[3 * inch, 2.5 * inch, 1.5 * inch])
    waste_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0fdf4')])
    ]))

    story.append(waste_table)
    story.append(Spacer(1, 0.3 * inch))

    # ==================== ENHANCED CHARTS SECTION ====================
    story.append(PageBreak())
    story.append(Paragraph("VISUAL ANALYTICS & INSIGHTS", heading_style))
    story.append(Spacer(1, 0.2 * inch))
    
    # Use enhanced PDF generator for high-quality charts
    from app.services.pdf_generator import PDFReportGenerator
    pdf_gen = PDFReportGenerator()
    
    # Emissions data for charts
    emissions_dict = {
        'scope1': report_data.summary_metrics.get('scope1_emissions', 0),
        'scope2': report_data.summary_metrics.get('scope2_emissions', 0),
        'scope3': report_data.summary_metrics.get('scope3_emissions', 0)
    }
    
    # 1. Scope Breakdown Pie Chart (Enhanced)
    story.append(Paragraph("📊 Emissions by Scope", heading_style))
    story.append(Spacer(1, 0.1 * inch))
    context_text = """
    <b>Understanding Scope Breakdown:</b> This chart shows the distribution of emissions across the three GHG Protocol scopes. 
    Scope 1 represents direct emissions from owned or controlled sources. Scope 2 covers indirect emissions from purchased energy. 
    Scope 3 includes all other indirect emissions in the value chain. Understanding this breakdown helps identify the most significant 
    emission sources and prioritize reduction efforts.
    """
    story.append(Paragraph(context_text, styles['Normal']))
    story.append(Spacer(1, 0.15 * inch))
    try:
        scope_pie = pdf_gen._create_scope_pie_chart(emissions_dict)
        if scope_pie:
            story.append(scope_pie)
    except Exception as e:
        story.append(Paragraph(f"<i>Chart generation error: {str(e)}</i>", styles['Normal']))
    story.append(Spacer(1, 0.3 * inch))
    
    # 2. Scope Breakdown Bar Chart
    story.append(Paragraph("📊 Scope Comparison (Bar Chart)", heading_style))
    story.append(Spacer(1, 0.1 * inch))
    try:
        scope_bar = pdf_gen._create_scope_bar_chart(emissions_dict)
        if scope_bar:
            story.append(scope_bar)
    except Exception as e:
        story.append(Paragraph(f"<i>Chart generation error: {str(e)}</i>", styles['Normal']))
    story.append(Spacer(1, 0.3 * inch))
    
    # 3. Trends Line Chart (Enhanced)
    if report_data.trends and len(report_data.trends) > 0:
        story.append(Paragraph("📈 Emissions Trend Over Time", heading_style))
        story.append(Spacer(1, 0.1 * inch))
        context_text = """
        <b>Trend Analysis:</b> This chart shows how emissions have changed over the reporting period. 
        Understanding trends helps identify patterns, seasonal variations, and the effectiveness of reduction initiatives. 
        A downward trend indicates successful emission reduction efforts, while an upward trend may signal increased activity 
        or the need for additional reduction measures.
        """
        story.append(Paragraph(context_text, styles['Normal']))
        story.append(Spacer(1, 0.15 * inch))
        try:
            trend_chart = pdf_gen._create_trend_line_chart(report_data.trends)
            if trend_chart:
                story.append(trend_chart)
        except Exception as e:
            story.append(Paragraph(f"<i>Trend chart generation error: {str(e)}</i>", styles['Normal']))
        story.append(Spacer(1, 0.3 * inch))
    
    # 4. Category Breakdown Chart
    if report_data.category_breakdown and len(report_data.category_breakdown) > 0:
        story.append(Paragraph("📦 Top Emission Categories", heading_style))
        story.append(Spacer(1, 0.1 * inch))
        context_text = """
        <b>Category Analysis:</b> This chart identifies the emission categories contributing most to your total footprint. 
        Categories are based on GHG Protocol classifications and help prioritize reduction efforts. 
        Focus on the top categories for maximum impact on overall emissions reduction.
        """
        story.append(Paragraph(context_text, styles['Normal']))
        story.append(Spacer(1, 0.15 * inch))
        try:
            category_chart = pdf_gen._create_category_bar_chart(report_data.category_breakdown)
            if category_chart:
                story.append(category_chart)
        except Exception as e:
            story.append(Paragraph(f"<i>Category chart generation error: {str(e)}</i>", styles['Normal']))
        story.append(Spacer(1, 0.3 * inch))
    
    # 5. Top Emitters Chart
    if report_data.activities and len(report_data.activities) > 0:
        story.append(Paragraph("🔥 Top Emission Sources", heading_style))
        story.append(Spacer(1, 0.1 * inch))
        context_text = """
        <b>Top Emitters Analysis:</b> This chart highlights the individual activities with the highest emissions. 
        These activities represent the greatest opportunities for emission reductions. 
        Consider implementing targeted reduction strategies for these high-impact activities.
        """
        story.append(Paragraph(context_text, styles['Normal']))
        story.append(Spacer(1, 0.15 * inch))
        try:
            top_emitters_data = [
                {
                    'activity': act.get('activity_name', 'N/A'),
                    'total_emissions_kg': act.get('emissions_kg', 0)
                }
                for act in report_data.activities[:8]
            ]
            top_emitters_chart = pdf_gen._create_top_emitters_chart(top_emitters_data)
            if top_emitters_chart:
                story.append(top_emitters_chart)
        except Exception as e:
            story.append(Paragraph(f"<i>Top emitters chart generation error: {str(e)}</i>", styles['Normal']))
        story.append(Spacer(1, 0.3 * inch))

    # ==================== PAGE 5: AI RECOMMENDATIONS ====================

    story.append(PageBreak())
    story.append(Paragraph("AI-POWERED RECOMMENDATIONS", heading_style))
    story.append(Spacer(1, 0.2 * inch))

    intro_text = """
    The following recommendations are generated using advanced AI analysis of your emission data, 
    industry best practices, and sustainability standards. Each recommendation includes priority level, 
    estimated impact, and implementation guidance.
    """
    story.append(Paragraph(intro_text, styles['Normal']))
    story.append(Spacer(1, 0.2 * inch))

    # Handle recommendations - show message if empty
    if report_data.recommendations and len(report_data.recommendations) > 0:
        # Handle recommendations - can be string or dict
        for idx, rec in enumerate(report_data.recommendations[:6], 1):
            if isinstance(rec, dict):
                rec_title = rec.get('title', f'Recommendation {idx}')
                rec_priority = rec.get('priority', 'Medium')
                rec_desc = rec.get('description', rec.get('detailed_analysis', 'No description available.'))
                rec_savings = rec.get('estimated_savings_kg', 0)
                
                # Priority color
                priority_color = '#dc2626' if rec_priority == 'High' else '#f59e0b' if rec_priority == 'Medium' else '#10b981'
                
                rec_text = f"""
                <b><font size=12 color={priority_color}>{idx}. {rec_title}</font></b><br/>
                <font size=9 color=#6b7280>Priority: {rec_priority} | Estimated Savings: {rec_savings / 1000:.2f} tons CO2e</font><br/><br/>
                {rec_desc[:400]}{'...' if len(rec_desc) > 400 else ''}
                """
                story.append(Paragraph(rec_text, styles['Normal']))
            else:
                story.append(Paragraph(f"{idx}. {rec}", styles['Normal']))
            story.append(Spacer(1, 0.15 * inch))
    else:
        story.append(Paragraph("<i>No AI recommendations available at this time. Recommendations will appear as more data is collected and analyzed.</i>", styles['Normal']))

    story.append(PageBreak())

    # ==================== PAGE 6: GOALS, ACTIVITIES & COMPLIANCE ====================

    # Goals section
    if report_data.goals:
        story.append(Paragraph("GOALS & TARGETS", heading_style))
        story.append(Spacer(1, 0.1 * inch))
        
        goals_data = [['Goal', 'Target', 'Current', 'Deadline', 'Status']]
        for goal in report_data.goals[:5]:
            goals_data.append([
                goal.get('title', 'N/A'),
                str(goal.get('target_value', 0)),
                str(goal.get('current_value', 0)),
                goal.get('deadline', 'N/A'),
                goal.get('status', 'N/A')
            ])
        
        goals_table = Table(goals_data, colWidths=[2 * inch, 1 * inch, 1 * inch, 1.2 * inch, 1.2 * inch])
        goals_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8b5cf6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#faf5ff')])
        ]))
        story.append(goals_table)
        story.append(Spacer(1, 0.3 * inch))

    # Top Activities
    story.append(Paragraph("TOP EMISSION ACTIVITIES", heading_style))
    story.append(Spacer(1, 0.1 * inch))
    if report_data.activities and len(report_data.activities) > 0:
        activities_data = [['Activity', 'Type', 'Emissions (kg CO2e)', 'Scope', 'Date']]
        for act in report_data.activities[:8]:
            activities_data.append([
                act.get('activity_name', 'N/A')[:30],
                act.get('activity_type', 'N/A')[:20],
                f"{act.get('emissions_kg', 0):,.2f}",
                act.get('scope', 'N/A'),
                act.get('date', 'N/A')
            ])
        
        activities_table = Table(activities_data, colWidths=[2 * inch, 1.5 * inch, 1.2 * inch, 0.8 * inch, 1 * inch])
        activities_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f59e0b')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fffbeb')])
        ]))
        story.append(activities_table)
    else:
        story.append(Paragraph("<i>No activities recorded for this reporting period.</i>", styles['Normal']))
    story.append(Spacer(1, 0.3 * inch))

    # Compliance & Standards
    story.append(Paragraph("COMPLIANCE & STANDARDS", heading_style))
    story.append(Spacer(1, 0.1 * inch))
    compliance = report_data.compliance_info
    compliance_text = f"""
    <b>Reporting Framework:</b> {compliance.get('reporting_framework', 'GHG Protocol Corporate Standard')}<br/>
    <b>GHG Protocol Compliant:</b> {'✓ Yes' if compliance.get('ghg_protocol_compliant') else '✗ No'}<br/>
    <b>ISO 14064 Compliant:</b> {'✓ Yes' if compliance.get('iso_14064_compliant') else '✗ No'}<br/>
    <b>GRI Standards Aligned:</b> {'✓ Yes' if compliance.get('gri_standards_aligned') else '✗ No'}<br/>
    <b>Verification Status:</b> {compliance.get('verification_status', 'Self-Reported')}<br/>
    <b>Data Quality:</b> {compliance.get('data_quality', 'High')}<br/>
    <b>Assurance Level:</b> {compliance.get('assurance_level', 'Limited Assurance')}
    """
    story.append(Paragraph(compliance_text, styles['Normal']))
    story.append(Spacer(1, 0.3 * inch))

    # Category breakdown if not already shown
    if report_data.category_breakdown and len(report_data.category_breakdown) > 0:
        story.append(Paragraph("EMISSIONS BY CATEGORY", heading_style))
        story.append(Spacer(1, 0.1 * inch))
        category_data = [['Category', 'Emissions (tons CO2e)', 'Percentage', 'Activities']]
        for cat in report_data.category_breakdown[:8]:
            category_data.append([
                cat.get('category', 'Uncategorized'),
                f"{cat.get('emissions_kg', 0) / 1000:.2f}",
                f"{cat.get('emissions_percent', 0):.1f}%",
                str(cat.get('activities_count', 0))
            ])

        category_table = Table(category_data, colWidths=[2.5 * inch, 1.5 * inch, 1.2 * inch, 1.2 * inch])
        category_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0fdf4')])
        ]))
        story.append(category_table)
        story.append(Spacer(1, 0.2 * inch))

    # ==================== FOOTER ====================

    story.append(Spacer(1, 0.5 * inch))
    footer_text = f"""
    <para align=center>
    <font size=9 color=#6b7280>
    <i>This report is generated by AI-Powered Carbon Accounting Platform<br/>
    For questions or additional information, contact {company.reporting_officer} at {company.email}<br/>
    Report generated on {report_data.generated_at.strftime('%Y-%m-%d %H:%M:%S')}</i>
    </font>
    </para>
    """
    story.append(Paragraph(footer_text, styles['Normal']))

    # Build PDF
    doc.build(story)
    buffer.seek(0)

    return buffer


# ==================== EXISTING ENDPOINTS (PRESERVED) ====================

@router.post("/generate", response_model=DetailedReport)
async def generate_report(
        config: ReportConfig,
        current_user: User = Depends(get_current_user),
        db: Session = Depends(get_db)
):
    """Generate a comprehensive emissions report (existing endpoint)"""

    # Validate date range
    if config.date_range.start_date > config.date_range.end_date:
        raise HTTPException(
            status_code=400,
            detail="Start date must be before end date"
        )

    # Calculate summary metrics
    summary = calculate_summary_metrics(
        db,
        current_user.company_id,
        config.date_range.start_date,
        config.date_range.end_date,
        config.scope,
        config.category
    )

    # Generate trend data
    trends = generate_trend_data(
        db,
        current_user.company_id,
        config.date_range.start_date,
        config.date_range.end_date,
        config.group_by,
        config.scope,
        config.category
    )

    # Generate category breakdown
    category_breakdown = generate_category_breakdown(
        db,
        current_user.company_id,
        config.date_range.start_date,
        config.date_range.end_date,
        config.scope
    )

    # Scope breakdown
    scope_query = db.query(
        EmissionActivity.scope_number,
        func.sum(EmissionActivity.emissions_kgco2e).label('total')
    ).filter(
        EmissionActivity.company_id == current_user.company_id,
        EmissionActivity.activity_date >= config.date_range.start_date,
        EmissionActivity.activity_date <= config.date_range.end_date
    ).group_by(EmissionActivity.scope_number).all()

    scope_breakdown = {f"Scope {scope}": round(total, 2) for scope, total in scope_query}

    # Top emission sources
    top_sources_query = db.query(
        EmissionActivity.activity_name,
        func.sum(EmissionActivity.emissions_kgco2e).label('total'),
        func.count(EmissionActivity.id).label('count')
    ).filter(
        EmissionActivity.company_id == current_user.company_id,
        EmissionActivity.activity_date >= config.date_range.start_date,
        EmissionActivity.activity_date <= config.date_range.end_date
    ).group_by(EmissionActivity.activity_name).order_by(func.sum(EmissionActivity.emissions_kgco2e).desc()).limit(
        10).all()

    top_sources = [
        {
            "activity": name,
            "total_emissions_kg": round(total, 2),
            "occurrences": count,
            "average_per_occurrence": round(total / count, 2) if count > 0 else 0
        }
        for name, total, count in top_sources_query
    ]

    # Goals progress
    goals = db.query(Goal).filter(Goal.user_id == current_user.id).all()
    goals_progress = [
        {
            "goal_title": goal.title,
            "progress_percent": round(goal.progress_percent, 2),
            "status": goal.status,
            "target_year": goal.target_year
        }
        for goal in goals
    ]

    # Recommendations summary
    recommendations_summary = {
        "high_priority_actions": 3,
        "potential_reduction_kg": round(summary.total_emissions_kg * 0.25, 2),
        "estimated_cost_savings": round(summary.total_emissions_kg * 0.05, 2)
    }

    return DetailedReport(
        summary=summary,
        trends=trends,
        category_breakdown=category_breakdown,
        scope_breakdown=scope_breakdown,
        top_emission_sources=top_sources,
        goals_progress=goals_progress,
        recommendations_summary=recommendations_summary
    )


@router.post("/generate-comprehensive-csv")
async def generate_comprehensive_csv_report(
        config: ReportConfig,
        current_user: User = Depends(get_current_user),
        db: Session = Depends(get_db)
):
    """Generate comprehensive CSV report with all activities and metrics"""
    print("🔍 CSV endpoint called - generate_comprehensive_csv_report")

    try:
        # Get company info
        company = db.query(Company).filter(Company.id == current_user.company_id).first()

        # Calculate all metrics
        summary = calculate_summary_metrics(
            db,
            current_user.company_id,
            config.date_range.start_date,
            config.date_range.end_date
        )

        water_data = get_water_metrics(db, current_user.company_id, config.date_range.start_date,
                                       config.date_range.end_date)
        waste_data = get_waste_metrics(db, current_user.company_id, config.date_range.start_date,
                                       config.date_range.end_date)
        scope3_data = get_scope3_metrics(db, current_user.company_id, config.date_range.start_date,
                                         config.date_range.end_date)

        trends_obj = generate_trend_data(
            db, current_user.company_id, config.date_range.start_date, config.date_range.end_date, config.group_by
        )

        # Get ALL activities (not just top 10)
        all_activities = []
        try:
            activities_query = db.query(EmissionActivity).filter(
                EmissionActivity.company_id == current_user.company_id,
                EmissionActivity.activity_date >= config.date_range.start_date,
                EmissionActivity.activity_date <= config.date_range.end_date
            ).order_by(EmissionActivity.activity_date.desc(), EmissionActivity.emissions_kgco2e.desc()).all()
            
            all_activities = [
                {
                    "activity_name": a.activity_name or a.activity_type or "Unnamed Activity",
                    "activity_type": a.activity_type or "N/A",
                    "description": a.description or "",
                    "quantity": a.quantity or 0,
                    "unit": a.unit or "N/A",
                    "emissions_kg": round(a.emissions_kgco2e, 2),
                    "emissions_tons": round(a.emissions_kgco2e / 1000, 3),
                    "scope": f"Scope {a.scope_number}",
                    "scope_number": a.scope_number or 0,
                    "category": a.category or "Uncategorized",
                    "subcategory": a.subcategory or "",
                    "location": a.location or "N/A",
                    "activity_date": a.activity_date.strftime("%Y-%m-%d") if a.activity_date else "N/A",
                    "reporting_period": a.reporting_period or "N/A",
                    "source_document": a.source_document or "",
                    "created_at": a.created_at.strftime("%Y-%m-%d %H:%M:%S") if a.created_at else "N/A"
                }
                for a in activities_query
            ]
        except Exception as e:
            print(f"⚠️ Error fetching all activities: {e}")

        # Get category breakdown
        category_breakdown_obj = generate_category_breakdown(
            db, current_user.company_id, config.date_range.start_date, config.date_range.end_date
        )
        category_breakdown = [
            {
                "category": c.category,
                "emissions_kg": c.emissions_kg,
                "emissions_tons": c.emissions_kg / 1000,
                "emissions_percent": c.emissions_percent,
                "activities_count": c.activities_count
            }
            for c in category_breakdown_obj
        ]

        # Fetch AI Recommendations
        ensure_ai_recommendation_schema()
        ai_recommendations = []
        try:
            cached_rec = db.query(AIRecommendation).filter(
                AIRecommendation.company_id == current_user.company_id,
                AIRecommendation.is_active == True
            ).order_by(AIRecommendation.generated_at.desc()).first()
            
            if cached_rec:
                try:
                    if hasattr(cached_rec, 'is_expired') and cached_rec.is_expired():
                        pass
                    else:
                        recs_json = cached_rec.recommendations_json
                        if recs_json:
                            if isinstance(recs_json, list):
                                ai_recommendations = recs_json
                            elif isinstance(recs_json, dict) and 'recommendations' in recs_json:
                                ai_recommendations = recs_json.get('recommendations', [])
                except Exception as rec_error:
                    print(f"⚠️ Error processing cached recommendations: {rec_error}")
        except Exception as e:
            print(f"⚠️ Error fetching AI recommendations: {e}")

        # Fetch Goals
        goals = []
        try:
            goals_query = db.query(Goal).filter(
                Goal.user_id == current_user.id
            ).all()
            goals = [
                {
                    "title": g.title,
                    "target_emissions": g.target_emissions,
                    "current_emissions": g.current_emissions or 0,
                    "target_year": g.target_year,
                    "status": g.status.replace("_", " ").title() if g.status else "On Track"
                }
                for g in goals_query
            ]
        except Exception as e:
            print(f"⚠️ Error fetching goals: {e}")

        # Create comprehensive CSV with all data
        # Combine all data into a single comprehensive CSV
        csv_rows = []
        
        # Header section
        csv_rows.append("=== SUSTAINABILITY REPORT ===")
        csv_rows.append(f"Company: {getattr(company, 'name', 'Your Company') if company else 'Your Company'}")
        csv_rows.append(f"Reporting Period: {config.date_range.start_date} to {config.date_range.end_date}")
        csv_rows.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        csv_rows.append("")
        
        # Summary Metrics
        csv_rows.append("=== SUMMARY METRICS ===")
        csv_rows.append("Metric,Value")
        csv_rows.append(f"Total Emissions (kg CO2e),{summary.total_emissions_kg}")
        csv_rows.append(f"Total Emissions (tons CO2e),{summary.total_emissions_tons}")
        csv_rows.append(f"Scope 1 Emissions,{summary.scope1_emissions}")
        csv_rows.append(f"Scope 2 Emissions,{summary.scope2_emissions}")
        csv_rows.append(f"Scope 3 Emissions,{summary.scope3_emissions}")
        csv_rows.append(f"Total Water Usage (m³),{water_data.get('total_usage', 0)}")
        csv_rows.append(f"Total Waste (kg),{waste_data.get('total_waste', 0)}")
        csv_rows.append(f"Recycled Waste (kg),{waste_data.get('recycled', 0)}")
        csv_rows.append(f"Recycling Rate (%),{waste_data.get('recycling_rate', 0)}")
        csv_rows.append(f"Total Activities Tracked,{summary.total_activities}")
        csv_rows.append("")
        
        # Emissions Breakdown
        csv_rows.append("=== EMISSIONS BREAKDOWN ===")
        csv_rows.append("Scope,Emissions (kg CO2e),Emissions (tons CO2e),Percentage (%)")
        total_kg = summary.total_emissions_kg if summary.total_emissions_kg > 0 else 1
        csv_rows.append(f"Scope 1,{summary.scope1_emissions},{summary.scope1_emissions / 1000},{round(summary.scope1_emissions / total_kg * 100, 1)}")
        csv_rows.append(f"Scope 2,{summary.scope2_emissions},{summary.scope2_emissions / 1000},{round(summary.scope2_emissions / total_kg * 100, 1)}")
        csv_rows.append(f"Scope 3,{summary.scope3_emissions},{summary.scope3_emissions / 1000},{round(summary.scope3_emissions / total_kg * 100, 1)}")
        csv_rows.append(f"TOTAL,{summary.total_emissions_kg},{summary.total_emissions_tons},100.0")
        csv_rows.append("")
        
        # Category Breakdown
        if category_breakdown:
            csv_rows.append("=== CATEGORY BREAKDOWN ===")
            csv_rows.append("Category,Emissions (kg CO2e),Emissions (tons CO2e),Percentage (%),Activities Count")
            for cat in category_breakdown:
                csv_rows.append(f"{cat['category']},{cat['emissions_kg']},{cat['emissions_tons']},{cat['emissions_percent']},{cat['activities_count']}")
            csv_rows.append("")
        
        # Trends
        if trends_obj:
            csv_rows.append("=== TRENDS ===")
            csv_rows.append("Period,Emissions (kg CO2e),Activities Count")
            for trend in trends_obj:
                csv_rows.append(f"{trend.period},{trend.emissions_kg},{trend.activities_count}")
            csv_rows.append("")
        
        # Water Usage
        csv_rows.append("=== WATER USAGE ===")
        csv_rows.append("Source,Usage (m³)")
        csv_rows.append(f"Municipal Water,{water_data.get('municipal_water', 0)}")
        csv_rows.append(f"Groundwater,{water_data.get('groundwater', 0)}")
        csv_rows.append(f"Rainwater,{water_data.get('rainwater', 0)}")
        csv_rows.append(f"TOTAL,{water_data.get('total_usage', 0)}")
        csv_rows.append("")
        
        # Waste Management
        csv_rows.append("=== WASTE MANAGEMENT ===")
        csv_rows.append("Disposal Method,Weight (kg)")
        csv_rows.append(f"Recycled,{waste_data.get('recycled', 0)}")
        csv_rows.append(f"Landfill,{waste_data.get('landfill', 0)}")
        csv_rows.append(f"Incinerated,{waste_data.get('incinerated', 0)}")
        csv_rows.append(f"TOTAL,{waste_data.get('total_waste', 0)}")
        csv_rows.append("")
        
        # All Activities
        csv_rows.append("=== ALL ACTIVITIES ===")
        if all_activities:
            csv_rows.append("Activity Name,Activity Type,Description,Quantity,Unit,Emissions (kg CO2e),Emissions (tons CO2e),Scope,Scope Number,Category,Subcategory,Location,Activity Date,Reporting Period,Source Document,Created At")
            for act in all_activities:
                csv_rows.append(f'"{act["activity_name"]}","{act["activity_type"]}","{act["description"]}",{act["quantity"]},"{act["unit"]}",{act["emissions_kg"]},{act["emissions_tons"]},"{act["scope"]}",{act["scope_number"]},"{act["category"]}","{act["subcategory"]}","{act["location"]}",{act["activity_date"]},"{act["reporting_period"]}","{act["source_document"]}",{act["created_at"]}')
        else:
            csv_rows.append("No activities found for this reporting period")
        csv_rows.append("")
        
        # AI Recommendations
        if ai_recommendations:
            csv_rows.append("=== AI RECOMMENDATIONS ===")
            csv_rows.append("Title,Priority,Estimated Savings (kg CO2e),Description")
            for rec in ai_recommendations:
                if isinstance(rec, dict):
                    title = rec.get('title', 'N/A').replace('"', '""')
                    desc = rec.get('description', rec.get('detailed_analysis', 'N/A')).replace('"', '""')
                    csv_rows.append(f'"{title}","{rec.get("priority", "Medium")}",{rec.get("estimated_savings_kg", 0)},"{desc}"')
                else:
                    rec_str = str(rec).replace('"', '""')
                    csv_rows.append(f'"{rec_str}","Medium",0,"{rec_str}"')
            csv_rows.append("")
        
        # Goals
        if goals:
            csv_rows.append("=== GOALS & TARGETS ===")
            csv_rows.append("Title,Target Emissions,Current Emissions,Target Year,Status")
            for goal in goals:
                title = goal.get('title', 'N/A').replace('"', '""')
                csv_rows.append(f'"{title}",{goal.get("target_emissions", 0)},{goal.get("current_emissions", 0)},{goal.get("target_year", "N/A")},"{goal.get("status", "On Track")}"')
            csv_rows.append("")

        # Combine all rows
        csv_content = "\n".join(csv_rows)

        filename = f"sustainability_report_{datetime.now().strftime('%Y%m%d')}.csv"

        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"❌ Error generating CSV report: {str(e)}")
        print(f"❌ Traceback: {error_details}")
        raise HTTPException(status_code=500, detail=f"Error generating CSV report: {str(e)}")


@router.post("/export/excel")
async def export_report_excel(
        config: ReportConfig,
        current_user: User = Depends(get_current_user),
        db: Session = Depends(get_db)
):
    """Export comprehensive report as Excel with multiple sheets"""

    # Generate report data
    report = await generate_report(config, current_user, db)

    # Create Excel file in memory
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Summary sheet
        summary_data = {
            "Metric": [
                "Total Emissions (kg)",
                "Total Emissions (tons)",
                "Scope 1 Emissions",
                "Scope 2 Emissions",
                "Scope 3 Emissions",
                "Top Category",
                "Top Category Emissions",
                "Average Daily Emissions",
                "Total Activities",
                "Comparison to Previous Period (%)"
            ],
            "Value": [
                report.summary.total_emissions_kg,
                report.summary.total_emissions_tons,
                report.summary.scope1_emissions,
                report.summary.scope2_emissions,
                report.summary.scope3_emissions,
                report.summary.top_category,
                report.summary.top_category_emissions,
                report.summary.average_daily_emissions,
                report.summary.total_activities,
                report.summary.comparison_to_previous_period
            ]
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)

        # Trends sheet
        trends_data = [{
            "Period": t.period,
            "Emissions (kg)": t.emissions_kg,
            "Activities": t.activities_count,
            "Avg per Activity": t.average_emission_per_activity
        } for t in report.trends]
        pd.DataFrame(trends_data).to_excel(writer, sheet_name='Trends', index=False)

        # Category breakdown sheet
        category_data = [{
            "Category": c.category,
            "Emissions (kg)": c.emissions_kg,
            "Percentage": c.emissions_percent,
            "Activities": c.activities_count
        } for c in report.category_breakdown]
        pd.DataFrame(category_data).to_excel(writer, sheet_name='Categories', index=False)

        # Top sources sheet
        pd.DataFrame(report.top_emission_sources).to_excel(writer, sheet_name='Top Sources', index=False)

        # Goals progress sheet
        if report.goals_progress:
            pd.DataFrame(report.goals_progress).to_excel(writer, sheet_name='Goals', index=False)

    output.seek(0)

    return StreamingResponse(
        io.BytesIO(output.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=emissions_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
        }
    )


@router.get("/dashboard")
async def get_dashboard_data(
        current_user: User = Depends(get_current_user),
        db: Session = Depends(get_db)
):
    """Get dashboard overview data"""

    # Current month data
    current_month_start = datetime.now().replace(day=1).date()
    current_month_end = date.today()

    # Last 30 days
    last_30_days_start = date.today() - timedelta(days=30)

    # Last 12 months
    last_12_months_start = date.today() - timedelta(days=365)

    # Current month summary
    current_month_summary = calculate_summary_metrics(
        db, current_user.company_id, current_month_start, current_month_end
    )

    # Last 30 days trend
    last_30_days_trend = generate_trend_data(
        db, current_user.company_id, last_30_days_start, date.today(), "day"
    )

    # Last 12 months trend
    last_12_months_trend = generate_trend_data(
        db, current_user.company_id, last_12_months_start, date.today(), "month"
    )

    # Category breakdown (last 30 days)
    category_breakdown = generate_category_breakdown(
        db, current_user.company_id, last_30_days_start, date.today()
    )

    # Active goals
    active_goals = db.query(Goal).filter(
        Goal.user_id == current_user.id,
        Goal.status.in_(["on_track", "at_risk"])
    ).all()

    goals_summary = [
        {
            "title": goal.title,
            "progress": round(goal.progress_percent, 1),
            "status": goal.status,
            "target_year": goal.target_year
        }
        for goal in active_goals[:3]
    ]

    # Recent activities
    recent_activities = db.query(EmissionActivity).filter(
        EmissionActivity.company_id == current_user.company_id
    ).order_by(EmissionActivity.activity_date.desc()).limit(10).all()

    recent_list = [
        {
            "id": r.id,
            "date": r.activity_date.isoformat() if r.activity_date else None,
            "activity": r.activity_name,
            "emissions_kg": round(r.emissions_kgco2e, 2),
            "category": r.category,
            "scope": r.scope_number
        }
        for r in recent_activities
    ]

    return {
        "current_month": {
            "total_emissions_kg": current_month_summary.total_emissions_kg,
            "comparison_percent": current_month_summary.comparison_to_previous_period,
            "activities_count": current_month_summary.total_activities,
            "average_daily": current_month_summary.average_daily_emissions
        },
        "last_30_days_trend": [
            {
                "date": t.period,
                "emissions": t.emissions_kg
            }
            for t in last_30_days_trend
        ],
        "last_12_months_trend": [
            {
                "month": t.period,
                "emissions": t.emissions_kg
            }
            for t in last_12_months_trend
        ],
        "category_breakdown": [
            {
                "category": c.category,
                "emissions": c.emissions_kg,
                "percent": c.emissions_percent
            }
            for c in category_breakdown[:5]
        ],
        "active_goals": goals_summary,
        "recent_activities": recent_list,
        "quick_stats": {
            "total_lifetime_emissions": db.query(func.sum(EmissionActivity.emissions_kgco2e)).filter(
                EmissionActivity.company_id == current_user.company_id
            ).scalar() or 0,
            "total_activities_tracked": db.query(func.count(EmissionActivity.id)).filter(
                EmissionActivity.company_id == current_user.company_id
            ).scalar() or 0,
            "active_goals_count": len(active_goals),
            "days_tracking": (date.today() - current_user.created_at.date()).days if current_user.created_at else 0
        }
    }


@router.get("/comparison/{year}")
async def get_year_comparison(
        year: int,
        current_user: User = Depends(get_current_user),
        db: Session = Depends(get_db)
):
    """Compare emissions across years"""

    current_year = datetime.now().year

    if year > current_year:
        raise HTTPException(status_code=400, detail="Cannot compare future years")

    years_to_compare = [year - 1, year, year + 1] if year < current_year else [year - 2, year - 1, year]
    years_to_compare = [y for y in years_to_compare if y <= current_year and y >= 2020]

    comparison_data = []

    for y in years_to_compare:
        start = date(y, 1, 1)
        end = date(y, 12, 31) if y < current_year else date.today()

        summary = calculate_summary_metrics(db, current_user.company_id, start, end)

        comparison_data.append({
            "year": y,
            "total_emissions_kg": summary.total_emissions_kg,
            "scope1": summary.scope1_emissions,
            "scope2": summary.scope2_emissions,
            "scope3": summary.scope3_emissions,
            "activities": summary.total_activities
        })

    return {
        "comparison": comparison_data,
        "analysis": {
            "best_year": min(comparison_data, key=lambda x: x["total_emissions_kg"])[
                "year"] if comparison_data else None,
            "worst_year": max(comparison_data, key=lambda x: x["total_emissions_kg"])[
                "year"] if comparison_data else None,
            "average_emissions": sum(d["total_emissions_kg"] for d in comparison_data) / len(
                comparison_data) if comparison_data else 0
        }
    }


# ==================== NEW COMPREHENSIVE REPORT ENDPOINTS ====================

@router.post("/generate-comprehensive-pdf")
async def generate_comprehensive_pdf_report(
        config: ReportConfig,
        current_user: User = Depends(get_current_user),
        db: Session = Depends(get_db)
):
    """Generate comprehensive PDF report with all metrics"""

    try:
        # Get company info
        company = db.query(Company).filter(Company.id == current_user.company_id).first()

        # Get company attributes with fallbacks
        company_name = getattr(company, 'name', 'Your Company') if company else "Your Company"
        company_address = getattr(company, 'address', None) if company else None
        company_industry = getattr(company, 'industry', None) if company else None
        company_phone = getattr(company, 'phone', None) if company else None
        company_website = getattr(company, 'website', None) if company else None

        company_profile = CompanyProfile(
            company_name=company_name,
            address=company_address if company_address else "Address not set",
            industry=company_industry if company_industry else "Industry not set",
            reporting_officer=getattr(current_user, 'full_name', current_user.email),
            email=current_user.email,
            phone=company_phone if company_phone else "N/A",
            website=company_website
        )

        # Calculate all metrics
        summary = calculate_summary_metrics(
            db,
            current_user.company_id,
            config.date_range.start_date,
            config.date_range.end_date
        )

        # Get water, waste, scope3 metrics
        water_data = get_water_metrics(db, current_user.company_id, config.date_range.start_date,
                                       config.date_range.end_date)
        waste_data = get_waste_metrics(db, current_user.company_id, config.date_range.start_date,
                                       config.date_range.end_date)
        scope3_data = get_scope3_metrics(db, current_user.company_id, config.date_range.start_date,
                                         config.date_range.end_date)

        # Get trends
        trends_obj = generate_trend_data(
            db, current_user.company_id, config.date_range.start_date, config.date_range.end_date, config.group_by
        )
        trends = [{"period": t.period, "emissions_kg": t.emissions_kg} for t in trends_obj]

        # Summary metrics dict
        summary_metrics = {
            "reporting_period_start": config.date_range.start_date.strftime("%Y-%m-%d"),
            "reporting_period_end": config.date_range.end_date.strftime("%Y-%m-%d"),
            "total_emissions_kg": summary.total_emissions_kg,
            "total_emissions_tons": summary.total_emissions_tons,
            "scope1_emissions": summary.scope1_emissions,
            "scope2_emissions": summary.scope2_emissions,
            "scope3_emissions": summary.scope3_emissions,
            "total_water_usage_m3": water_data.get('total_usage', 0),
            "total_waste_kg": waste_data.get('total_waste', 0),
            "recycled_waste_kg": waste_data.get('recycled', 0),
            "recycling_rate": waste_data.get('recycling_rate', 0),
            "total_activities": summary.total_activities,
            "reporting_period_days": (config.date_range.end_date - config.date_range.start_date).days + 1
        }

        # Emissions data for charts
        emissions_data = {
            "scope1_emissions": summary.scope1_emissions,
            "scope2_emissions": summary.scope2_emissions,
            "scope3_emissions": summary.scope3_emissions,
            "total_emissions_kg": summary.total_emissions_kg
        }

        # Fetch AI Recommendations from database
        ensure_ai_recommendation_schema()
        ai_recommendations = []
        try:
            cached_rec = db.query(AIRecommendation).filter(
                AIRecommendation.company_id == current_user.company_id,
                AIRecommendation.is_active == True
            ).order_by(AIRecommendation.generated_at.desc()).first()
            
            if cached_rec:
                # Check if expired (if method exists)
                try:
                    if hasattr(cached_rec, 'is_expired') and cached_rec.is_expired():
                        pass  # Skip expired recommendations
                    else:
                        recs_json = cached_rec.recommendations_json
                        if recs_json:
                            if isinstance(recs_json, list):
                                ai_recommendations = recs_json[:6]  # Get up to 6 recommendations
                            elif isinstance(recs_json, dict) and 'recommendations' in recs_json:
                                # Handle case where recommendations_json is a dict with 'recommendations' key
                                ai_recommendations = recs_json.get('recommendations', [])[:6]
                except Exception as rec_error:
                    print(f"⚠️ Error processing cached recommendations: {rec_error}")
        except Exception as e:
            print(f"⚠️ Error fetching AI recommendations: {e}")
            import traceback
            print(traceback.format_exc())
        
        # If no AI recommendations, return empty list (no fallback data)
        # The PDF will show "No recommendations available" if empty

        # Fetch Goals from database
        goals = []
        try:
            goals_query = db.query(Goal).filter(
                Goal.user_id == current_user.id
            ).all()
            goals = [
                {
                    "title": g.title,
                    "target_value": g.target_emissions,
                    "current_value": g.current_emissions or 0,
                    "deadline": f"{g.target_year}-12-31",
                    "status": g.status.replace("_", " ").title() if g.status else "On Track"
                }
                for g in goals_query
            ]
        except Exception as e:
            print(f"⚠️ Error fetching goals: {e}")
            import traceback
            print(traceback.format_exc())

        # Get top activities
        top_activities = []
        try:
            activities_query = db.query(EmissionActivity).filter(
                EmissionActivity.company_id == current_user.company_id,
                EmissionActivity.activity_date >= config.date_range.start_date,
                EmissionActivity.activity_date <= config.date_range.end_date
            ).order_by(EmissionActivity.emissions_kgco2e.desc()).limit(10).all()
            
            top_activities = [
                {
                    "activity_name": a.activity_name or a.activity_type or "Unnamed Activity",
                    "activity_type": a.activity_type or "N/A",
                    "emissions_kg": round(a.emissions_kgco2e, 2),
                    "scope": f"Scope {a.scope_number}",
                    "date": a.activity_date.strftime("%Y-%m-%d") if a.activity_date else "N/A"
                }
                for a in activities_query
            ]
        except Exception as e:
            print(f"⚠️ Error fetching activities: {e}")

        # Get category breakdown
        category_breakdown_obj = generate_category_breakdown(
            db, current_user.company_id, config.date_range.start_date, config.date_range.end_date
        )
        category_breakdown = [
            {
                "category": c.category,
                "emissions_kg": c.emissions_kg,
                "emissions_percent": c.emissions_percent,
                "activities_count": c.activities_count
            }
            for c in category_breakdown_obj
        ]

        # Compliance information
        compliance_info = {
            "ghg_protocol_compliant": True,
            "iso_14064_compliant": True,
            "gri_standards_aligned": True,
            "reporting_framework": "GHG Protocol Corporate Standard",
            "verification_status": "Self-Reported",
            "data_quality": "High",
            "assurance_level": "Limited Assurance"
        }

        # Create comprehensive report data
        report_data = ComprehensiveReportData(
            company_profile=company_profile,
            summary_metrics=summary_metrics,
            emissions_data=emissions_data,
            water_data=water_data,
            waste_data=waste_data,
            scope3_data=scope3_data,
            trends=trends,
            recommendations=ai_recommendations,
            goals=goals,
            activities=top_activities,
            category_breakdown=category_breakdown,
            compliance_info=compliance_info,
            generated_at=datetime.now()
        )

        # Generate PDF
        pdf_buffer = generate_pdf_report(report_data)

        filename = f"sustainability_report_{company_profile.company_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"

        # Reset buffer position to start
        pdf_buffer.seek(0)
        
        # Return the buffer content as bytes
        return Response(
            content=pdf_buffer.getvalue(),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"❌ Error generating PDF report: {str(e)}")
        print(f"❌ Traceback: {error_details}")
        raise HTTPException(status_code=500, detail=f"Error generating PDF report: {str(e)}")


@router.post("/generate-comprehensive-excel")
async def generate_comprehensive_excel_report(
        config: ReportConfig,
        current_user: User = Depends(get_current_user),
        db: Session = Depends(get_db)
):
    """Generate comprehensive Excel report with all metrics"""

    try:
        # Get company info
        company = db.query(Company).filter(Company.id == current_user.company_id).first()

        # Calculate all metrics
        summary = calculate_summary_metrics(
            db,
            current_user.company_id,
            config.date_range.start_date,
            config.date_range.end_date
        )

        water_data = get_water_metrics(db, current_user.company_id, config.date_range.start_date,
                                       config.date_range.end_date)
        waste_data = get_waste_metrics(db, current_user.company_id, config.date_range.start_date,
                                       config.date_range.end_date)
        scope3_data = get_scope3_metrics(db, current_user.company_id, config.date_range.start_date,
                                         config.date_range.end_date)

        trends_obj = generate_trend_data(
            db, current_user.company_id, config.date_range.start_date, config.date_range.end_date, config.group_by
        )

        # Get ALL activities (not just top 10)
        all_activities = []
        try:
            activities_query = db.query(EmissionActivity).filter(
                EmissionActivity.company_id == current_user.company_id,
                EmissionActivity.activity_date >= config.date_range.start_date,
                EmissionActivity.activity_date <= config.date_range.end_date
            ).order_by(EmissionActivity.activity_date.desc(), EmissionActivity.emissions_kgco2e.desc()).all()
            
            all_activities = [
                {
                    "activity_name": a.activity_name or a.activity_type or "Unnamed Activity",
                    "activity_type": a.activity_type or "N/A",
                    "description": a.description or "",
                    "quantity": a.quantity or 0,
                    "unit": a.unit or "N/A",
                    "emissions_kg": round(a.emissions_kgco2e, 2),
                    "emissions_tons": round(a.emissions_kgco2e / 1000, 3),
                    "scope": f"Scope {a.scope_number}",
                    "scope_number": a.scope_number or 0,
                    "category": a.category or "Uncategorized",
                    "subcategory": a.subcategory or "",
                    "location": a.location or "N/A",
                    "activity_date": a.activity_date.strftime("%Y-%m-%d") if a.activity_date else "N/A",
                    "reporting_period": a.reporting_period or "N/A",
                    "source_document": a.source_document or "",
                    "created_at": a.created_at.strftime("%Y-%m-%d %H:%M:%S") if a.created_at else "N/A"
                }
                for a in activities_query
            ]
        except Exception as e:
            print(f"⚠️ Error fetching all activities: {e}")
            import traceback
            print(traceback.format_exc())

        # Get category breakdown
        category_breakdown_obj = generate_category_breakdown(
            db, current_user.company_id, config.date_range.start_date, config.date_range.end_date
        )
        category_breakdown = [
            {
                "category": c.category,
                "emissions_kg": c.emissions_kg,
                "emissions_tons": c.emissions_kg / 1000,
                "emissions_percent": c.emissions_percent,
                "activities_count": c.activities_count
            }
            for c in category_breakdown_obj
        ]

        # Fetch AI Recommendations
        ensure_ai_recommendation_schema()
        ai_recommendations = []
        try:
            cached_rec = db.query(AIRecommendation).filter(
                AIRecommendation.company_id == current_user.company_id,
                AIRecommendation.is_active == True
            ).order_by(AIRecommendation.generated_at.desc()).first()
            
            if cached_rec:
                try:
                    if hasattr(cached_rec, 'is_expired') and cached_rec.is_expired():
                        pass
                    else:
                        recs_json = cached_rec.recommendations_json
                        if recs_json:
                            if isinstance(recs_json, list):
                                ai_recommendations = recs_json
                            elif isinstance(recs_json, dict) and 'recommendations' in recs_json:
                                ai_recommendations = recs_json.get('recommendations', [])
                except Exception as rec_error:
                    print(f"⚠️ Error processing cached recommendations: {rec_error}")
        except Exception as e:
            print(f"⚠️ Error fetching AI recommendations: {e}")

        # Fetch Goals
        goals = []
        try:
            goals_query = db.query(Goal).filter(
                Goal.user_id == current_user.id
            ).all()
            goals = [
                {
                    "title": g.title,
                    "target_emissions": g.target_emissions,
                    "current_emissions": g.current_emissions or 0,
                    "target_year": g.target_year,
                    "status": g.status.replace("_", " ").title() if g.status else "On Track"
                }
                for g in goals_query
            ]
        except Exception as e:
            print(f"⚠️ Error fetching goals: {e}")

        # Create Excel file in memory
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:

            # Summary sheet
            summary_data = {
                'Metric': [
                    'Reporting Period Start',
                    'Reporting Period End',
                    'Total Emissions (kg CO2e)',
                    'Total Emissions (tons CO2e)',
                    'Scope 1 Emissions',
                    'Scope 2 Emissions',
                    'Scope 3 Emissions',
                    'Total Water Usage (m³)',
                    'Total Waste (kg)',
                    'Recycled Waste (kg)',
                    'Recycling Rate (%)',
                    'Total Activities Tracked'
                ],
                'Value': [
                    config.date_range.start_date.strftime("%Y-%m-%d"),
                    config.date_range.end_date.strftime("%Y-%m-%d"),
                    summary.total_emissions_kg,
                    summary.total_emissions_tons,
                    summary.scope1_emissions,
                    summary.scope2_emissions,
                    summary.scope3_emissions,
                    water_data.get('total_usage', 0),
                    waste_data.get('total_waste', 0),
                    waste_data.get('recycled', 0),
                    waste_data.get('recycling_rate', 0),
                    summary.total_activities
                ]
            }
            pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)

            # Emissions breakdown
            total_kg = summary.total_emissions_kg if summary.total_emissions_kg > 0 else 1
            emissions_breakdown = {
                'Scope': ['Scope 1', 'Scope 2', 'Scope 3', 'TOTAL'],
                'Emissions (kg CO2e)': [
                    summary.scope1_emissions,
                    summary.scope2_emissions,
                    summary.scope3_emissions,
                    summary.total_emissions_kg
                ],
                'Emissions (tons CO2e)': [
                    summary.scope1_emissions / 1000,
                    summary.scope2_emissions / 1000,
                    summary.scope3_emissions / 1000,
                    summary.total_emissions_tons
                ],
                'Percentage (%)': [
                    round(summary.scope1_emissions / total_kg * 100, 1),
                    round(summary.scope2_emissions / total_kg * 100, 1),
                    round(summary.scope3_emissions / total_kg * 100, 1),
                    100.0
                ]
            }
            pd.DataFrame(emissions_breakdown).to_excel(writer, sheet_name='Emissions Breakdown', index=False)

            # Water usage
            water_usage = {
                'Source': ['Municipal Water', 'Groundwater', 'Rainwater', 'TOTAL'],
                'Usage (m³)': [
                    water_data.get('municipal_water', 0),
                    water_data.get('groundwater', 0),
                    water_data.get('rainwater', 0),
                    water_data.get('total_usage', 0)
                ]
            }
            pd.DataFrame(water_usage).to_excel(writer, sheet_name='Water Usage', index=False)

            # Waste management
            waste_management = {
                'Disposal Method': ['Recycled', 'Landfill', 'Incinerated', 'TOTAL'],
                'Weight (kg)': [
                    waste_data.get('recycled', 0),
                    waste_data.get('landfill', 0),
                    waste_data.get('incinerated', 0),
                    waste_data.get('total_waste', 0)
                ]
            }
            pd.DataFrame(waste_management).to_excel(writer, sheet_name='Waste Management', index=False)

            # Trends
            trends_data = [{
                'Period': t.period,
                'Emissions (kg CO2e)': t.emissions_kg,
                'Activities': t.activities_count
            } for t in trends_obj]
            if trends_data:
                pd.DataFrame(trends_data).to_excel(writer, sheet_name='Trends', index=False)

            # ALL Activities - Detailed Sheet
            if all_activities:
                activities_df = pd.DataFrame(all_activities)
                activities_df.to_excel(writer, sheet_name='All Activities', index=False)
            else:
                # Create empty sheet with headers
                empty_activities = pd.DataFrame(columns=[
                    'activity_name', 'activity_type', 'description', 'quantity', 'unit',
                    'emissions_kg', 'emissions_tons', 'scope', 'scope_number', 'category',
                    'subcategory', 'location', 'activity_date', 'reporting_period',
                    'source_document', 'created_at'
                ])
                empty_activities.to_excel(writer, sheet_name='All Activities', index=False)

            # Category Breakdown Sheet
            if category_breakdown:
                category_df = pd.DataFrame(category_breakdown)
                category_df.to_excel(writer, sheet_name='Category Breakdown', index=False)
            else:
                empty_category = pd.DataFrame(columns=['category', 'emissions_kg', 'emissions_tons', 'emissions_percent', 'activities_count'])
                empty_category.to_excel(writer, sheet_name='Category Breakdown', index=False)

            # AI Recommendations Sheet
            if ai_recommendations:
                recommendations_data = []
                for rec in ai_recommendations:
                    if isinstance(rec, dict):
                        recommendations_data.append({
                            'Title': rec.get('title', 'N/A'),
                            'Priority': rec.get('priority', 'Medium'),
                            'Estimated Savings (kg CO2e)': rec.get('estimated_savings_kg', 0),
                            'Estimated Savings (tons CO2e)': round(rec.get('estimated_savings_kg', 0) / 1000, 2),
                            'Description': rec.get('description', rec.get('detailed_analysis', 'N/A'))
                        })
                    else:
                        recommendations_data.append({
                            'Title': str(rec),
                            'Priority': 'Medium',
                            'Estimated Savings (kg CO2e)': 0,
                            'Estimated Savings (tons CO2e)': 0,
                            'Description': str(rec)
                        })
                rec_df = pd.DataFrame(recommendations_data)
                rec_df.to_excel(writer, sheet_name='AI Recommendations', index=False)
            else:
                empty_rec = pd.DataFrame(columns=['Title', 'Priority', 'Estimated Savings (kg CO2e)', 'Estimated Savings (tons CO2e)', 'Description'])
                empty_rec.to_excel(writer, sheet_name='AI Recommendations', index=False)

            # Goals Sheet
            if goals:
                goals_df = pd.DataFrame(goals)
                goals_df.to_excel(writer, sheet_name='Goals & Targets', index=False)
            else:
                empty_goals = pd.DataFrame(columns=['title', 'target_emissions', 'current_emissions', 'target_year', 'status'])
                empty_goals.to_excel(writer, sheet_name='Goals & Targets', index=False)

            # Company info
            company_info = {
                'Field': ['Company Name', 'Reporting Officer', 'Email', 'Report Generated'],
                'Value': [
                    getattr(company, 'name', 'Your Company') if company else "Your Company",
                    getattr(current_user, 'full_name', current_user.email),
                    current_user.email,
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ]
            }
            pd.DataFrame(company_info).to_excel(writer, sheet_name='Company Info', index=False)

            # Get the workbook to add charts
            workbook = writer.book

            # Add charts to appropriate sheets
            try:
                # Chart 1: Emissions by Scope (Pie Chart) - Add to Summary sheet
                summary_sheet = workbook['Summary']
                
                # Create pie chart for emissions breakdown
                pie_chart = PieChart()
                pie_chart.title = "Emissions by Scope"
                pie_chart.dataLabels = DataLabelList()
                pie_chart.dataLabels.showPercent = True
                
                # Data for pie chart (from Emissions Breakdown sheet)
                emissions_sheet = workbook['Emissions Breakdown']
                # Reference to emissions values (skip header row)
                data = Reference(emissions_sheet, min_col=2, min_row=2, max_row=4)
                labels = Reference(emissions_sheet, min_col=1, min_row=2, max_row=4)
                pie_chart.add_data(data, titles_from_data=False)
                pie_chart.set_categories(labels)
                pie_chart.width = 10
                pie_chart.height = 7
                summary_sheet.add_chart(pie_chart, "D2")

                # Chart 2: Emissions Trend (Line Chart) - Add to Trends sheet if data exists
                if trends_data:
                    trends_sheet = workbook['Trends']
                    line_chart = LineChart()
                    line_chart.title = "Emissions Trend Over Time"
                    line_chart.style = 10
                    line_chart.y_axis.title = 'Emissions (kg CO2e)'
                    line_chart.x_axis.title = 'Period'
                    
                    data = Reference(trends_sheet, min_col=2, min_row=2, max_row=len(trends_data) + 1)
                    categories = Reference(trends_sheet, min_col=1, min_row=2, max_row=len(trends_data) + 1)
                    line_chart.add_data(data, titles_from_data=False)
                    line_chart.set_categories(categories)
                    line_chart.width = 15
                    line_chart.height = 7
                    trends_sheet.add_chart(line_chart, "E2")

                # Chart 3: Category Breakdown (Bar Chart) - Add to Category Breakdown sheet
                if category_breakdown:
                    category_sheet = workbook['Category Breakdown']
                    bar_chart = BarChart()
                    bar_chart.type = "col"
                    bar_chart.style = 10
                    bar_chart.title = "Emissions by Category"
                    bar_chart.y_axis.title = 'Emissions (tons CO2e)'
                    bar_chart.x_axis.title = 'Category'
                    
                    data = Reference(category_sheet, min_col=3, min_row=2, max_row=min(len(category_breakdown) + 1, 20))
                    categories = Reference(category_sheet, min_col=1, min_row=2, max_row=min(len(category_breakdown) + 1, 20))
                    bar_chart.add_data(data, titles_from_data=False)
                    bar_chart.set_categories(categories)
                    bar_chart.width = 15
                    bar_chart.height = 7
                    category_sheet.add_chart(bar_chart, "F2")

                # Chart 4: Scope Breakdown (Bar Chart) - Add to Emissions Breakdown sheet
                scope_bar = BarChart()
                scope_bar.type = "col"
                scope_bar.style = 10
                scope_bar.title = "Emissions by Scope"
                scope_bar.y_axis.title = 'Emissions (kg CO2e)'
                scope_bar.x_axis.title = 'Scope'
                
                data = Reference(emissions_sheet, min_col=2, min_row=2, max_row=4)
                categories = Reference(emissions_sheet, min_col=1, min_row=2, max_row=4)
                scope_bar.add_data(data, titles_from_data=False)
                scope_bar.set_categories(categories)
                scope_bar.width = 10
                scope_bar.height = 7
                emissions_sheet.add_chart(scope_bar, "E2")

            except Exception as chart_error:
                print(f"⚠️ Error adding charts to Excel: {chart_error}")
                import traceback
                print(traceback.format_exc())

            # Style the sheets
            try:
                from openpyxl.styles import Font, PatternFill, Alignment
                
                # Style header rows
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    if sheet.max_row > 0:
                        # Style header row
                        for cell in sheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        # Auto-adjust column widths
                        for column in sheet.columns:
                            max_length = 0
                            column_letter = get_column_letter(column[0].column)
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)
                            sheet.column_dimensions[column_letter].width = adjusted_width
            except Exception as style_error:
                print(f"⚠️ Error styling Excel sheets: {style_error}")

        output.seek(0)

        filename = f"sustainability_report_{datetime.now().strftime('%Y%m%d')}.xlsx"

        return StreamingResponse(
            io.BytesIO(output.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating Excel report: {str(e)}")


@router.get("/test-csv-endpoint")
async def test_csv_endpoint():
    """Test endpoint to verify CSV route is registered"""
    return {"message": "CSV endpoint is accessible", "status": "ok"}

@router.get("/company-profile")
async def get_company_profile(
        current_user: User = Depends(get_current_user),
        db: Session = Depends(get_db)
):
    """Get company profile"""

    try:
        company = db.query(Company).filter(Company.id == current_user.company_id).first()

        if not company:
            # Return default profile if company not found
            profile = CompanyProfile(
                company_name="Your Company",
                address="Company Address",
                industry="Industry",
                reporting_officer=getattr(current_user, 'full_name', current_user.email),
                email=current_user.email,
                phone="N/A",
                website=None
            )
            return profile

        profile = CompanyProfile(
            company_name=getattr(company, 'name', 'Your Company'),
            address=getattr(company, 'address', 'Address not set'),
            industry=getattr(company, 'industry', 'Industry not set'),
            reporting_officer=getattr(current_user, 'full_name', current_user.email),
            email=current_user.email,
            phone=getattr(company, 'phone', 'N/A'),
            website=getattr(company, 'website', None)
        )

        return profile

    except Exception as e:
        # Fallback to default profile on any error
        return CompanyProfile(
            company_name="Your Company",
            address="Company Address",
            industry="Industry",
            reporting_officer=getattr(current_user, 'full_name',
                                      current_user.email) if current_user else "Reporting Officer",
            email=current_user.email if current_user else "email@company.com",
            phone="N/A",
            website=None
        )

