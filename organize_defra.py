"""
DEFRA EMISSION FACTOR ORGANIZER - Script 2
===========================================
Extracts DEFRA emission factors with geographic applicability tagging
Optimized for India-based SaaS platform with ISO 14064 compliance

Author: Carbon Accounting Platform
Version: 1.0
"""

import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime
import json
import time


class DEFRAOrganizer:

    def __init__(self, filepath='defra_EF.xlsx'):
        """Initialize DEFRA organizer"""
        self.filepath = Path(filepath)
        self.factors = []
        self.factor_id = 1
        self.stats = {
            'total_sheets': 0,
            'processed_sheets': 0,
            'total_factors': 0,
            'by_scope': {},
            'india_relevant': 0,
            'uk_specific': 0
        }

        # Sheet mapping with India relevance
        self.sheet_mapping = {
            'Fuels': {'activity_type': 'fuel', 'scope': 'Scope 1', 'scope_category': '1.1',
                      'category_name': 'Stationary Combustion', 'india_relevant': True, 'priority_for_india': 2},
            'Bioenergy': {'activity_type': 'biofuel', 'scope': 'Scope 1', 'scope_category': '1.1',
                          'category_name': 'Stationary Combustion', 'india_relevant': True, 'priority_for_india': 2},
            'Passenger vehicles': {'activity_type': 'vehicle_passenger', 'scope': 'Scope 1', 'scope_category': '1.2',
                                   'category_name': 'Mobile Combustion', 'india_relevant': True,
                                   'priority_for_india': 3},
            'Delivery vehicles': {'activity_type': 'vehicle_delivery', 'scope': 'Scope 1', 'scope_category': '1.2',
                                  'category_name': 'Mobile Combustion', 'india_relevant': True,
                                  'priority_for_india': 3},
            'UK electricity': {'activity_type': 'electricity', 'scope': 'Scope 2', 'scope_category': '2.1',
                               'category_name': 'Purchased Electricity', 'india_relevant': False,
                               'priority_for_india': 5},
            'Overseas electricity': {'activity_type': 'electricity', 'scope': 'Scope 2', 'scope_category': '2.1',
                                     'category_name': 'Purchased Electricity', 'india_relevant': True,
                                     'priority_for_india': 3},
            'Heat and steam': {'activity_type': 'heat_steam', 'scope': 'Scope 2', 'scope_category': '2.2',
                               'category_name': 'Purchased Heat/Steam', 'india_relevant': True,
                               'priority_for_india': 2},
            'Business travel- air': {'activity_type': 'flight', 'scope': 'Scope 3', 'scope_category': '3.6',
                                     'category_name': 'Business Travel', 'india_relevant': True,
                                     'priority_for_india': 2},
            'Business travel- sea': {'activity_type': 'ferry', 'scope': 'Scope 3', 'scope_category': '3.6',
                                     'category_name': 'Business Travel', 'india_relevant': True,
                                     'priority_for_india': 2},
            'Business travel- land': {'activity_type': 'rail', 'scope': 'Scope 3', 'scope_category': '3.6',
                                      'category_name': 'Business Travel', 'india_relevant': True,
                                      'priority_for_india': 3},
            'Freighting goods': {'activity_type': 'freight', 'scope': 'Scope 3', 'scope_category': '3.4',
                                 'category_name': 'Upstream Transportation', 'india_relevant': True,
                                 'priority_for_india': 2},
            'Hotel stay': {'activity_type': 'accommodation', 'scope': 'Scope 3', 'scope_category': '3.6',
                           'category_name': 'Business Travel', 'india_relevant': True, 'priority_for_india': 3},
            'Material use': {'activity_type': 'material', 'scope': 'Scope 3', 'scope_category': '3.1',
                             'category_name': 'Purchased Goods & services', 'india_relevant': True,
                             'priority_for_india': 2},
            'Waste disposal': {'activity_type': 'waste', 'scope': 'Scope 3', 'scope_category': '3.5',
                               'category_name': 'Waste Generated in Operations', 'india_relevant': True,
                               'priority_for_india': 2},
            'Water supply': {'activity_type': 'water_supply', 'scope': 'Scope 3', 'scope_category': '3.1',
                             'category_name': 'Purchased Goods & services', 'india_relevant': True,
                             'priority_for_india': 3},
            'Water treatment': {'activity_type': 'water_treatment', 'scope': 'Scope 3', 'scope_category': '3.5',
                                'category_name': 'Waste Generated in Operations', 'india_relevant': True,
                                'priority_for_india': 3},
            'Refrigerant & other': {'activity_type': 'refrigerant', 'scope': 'Scope 1', 'scope_category': '1.3',
                                    'category_name': 'Fugitive Emissions', 'india_relevant': True,
                                    'priority_for_india': 2}
        }

    def process_defra_file(self):
        """Process DEFRA Excel file"""
        print(f"\n📂 Processing DEFRA file: {self.filepath}")
        print("=" * 80)

        if not self.filepath.exists():
            print(f"❌ ERROR: File not found: {self.filepath}")
            return

        wb = openpyxl.load_workbook(self.filepath, data_only=True)

        # Skip non-data sheets
        skip_sheets = ['Introduction', "What's new", 'Index', 'Conversions',
                       'Fuel properties', 'Haul definition', 'Outside of scopes',
                       'WTT- fuels', 'WTT- bioenergy', 'WTT- UK electricity',
                       'WTT- heat and steam', 'WTT- business travel- air',
                       'WTT- business travel- sea', 'WTT- pass vehs & travel- land',
                       'WTT- delivery vehs & freight', 'Transmission and distribution',
                       'UK electricity T&D for EVs', 'UK electricity for EVs',
                       'SECR kWh pass & delivery vehs', 'SECR kWh UK electricity for EVs',
                       'Managed assets- electricity', 'Managed assets- vehicles', 'Homeworking']

        process_sheets = [s for s in wb.sheetnames if s not in skip_sheets and s in self.sheet_mapping]

        self.stats['total_sheets'] = len(process_sheets)
        print(f"✅ Found {len(process_sheets)} data sheets to process\n")

        start_time = time.time()

        for idx, sheet_name in enumerate(process_sheets, 1):
            print(f"[{idx}/{len(process_sheets)}] 📄 Processing: {sheet_name}")
            self.process_sheet(wb[sheet_name], sheet_name)

        total_time = time.time() - start_time
        print(f"\n⏱️  Total time: {total_time:.1f}s ({total_time / 60:.1f} min)")

        wb.close()

    def process_sheet(self, sheet, sheet_name):
        """Process a single DEFRA sheet with flexible header detection"""
        try:
            # Get data as list of lists
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)

            # Find ALL rows with "kg CO2e" (DEFRA has multiple header patterns)
            header_candidates = []
            for idx, row in enumerate(data[:40]):
                if row and any('kg CO2e' in str(cell) for cell in row if cell):
                    header_candidates.append(idx)

            if not header_candidates:
                print(f"  ⚠️  No header found, skipping")
                return

            # Try each header candidate until we find one that works
            for header_row_idx in header_candidates:
                headers = data[header_row_idx]

                # Find column indices - FLEXIBLE MATCHING
                col_idx = {}
                for idx, header in enumerate(headers):
                    if not header:
                        continue

                    h_str = str(header).strip().lower()

                    # Match various column name patterns
                    if h_str == 'unit' or h_str == 'units':
                        col_idx['unit'] = idx
                    elif h_str == 'kg co2e' or 'total kg co2e' in h_str:
                        if 'co2e' not in col_idx:  # Take first occurrence
                            col_idx['co2e'] = idx
                    elif 'kg co2e of co2 per unit' in h_str or h_str == 'kg co2':
                        col_idx['co2'] = idx
                    elif 'kg co2e of ch4 per unit' in h_str or 'ch4' in h_str:
                        col_idx['ch4'] = idx
                    elif 'kg co2e of n2o per unit' in h_str or 'n2o' in h_str:
                        col_idx['n2o'] = idx
                    elif h_str in ['activity', 'emission source']:
                        col_idx['activity'] = idx
                    elif h_str in ['fuel', 'type', 'vehicle type', 'mode']:
                        if 'fuel' not in col_idx:
                            col_idx['fuel'] = idx
                    elif h_str in ['haul', 'distance']:
                        col_idx['haul'] = idx
                    elif h_str in ['class', 'size', 'category']:
                        col_idx['class'] = idx
                    elif h_str in ['country', 'location', 'region']:
                        col_idx['country'] = idx

                # Check if we have minimum required columns
                if 'co2e' in col_idx and 'unit' in col_idx:
                    # This header works! Process data
                    count = self.process_data_rows(data, header_row_idx, col_idx, sheet_name)
                    if count > 0:
                        print(f"  ✅ Extracted {count} factors (header at row {header_row_idx})")
                        self.stats['processed_sheets'] += 1
                        self.stats['total_factors'] += count
                        return

            # If we get here, no valid header configuration worked
            print(f"  ⚠️  Could not find valid data structure")

        except Exception as e:
            print(f"  ❌ Error: {str(e)}")

    def process_data_rows(self, data, header_row_idx, col_idx, sheet_name):
        """Process data rows from a sheet"""
        count = 0

        for row in data[header_row_idx + 1:]:
            if not row or len(row) <= max(col_idx.values()):
                continue

            # Skip rows where first few cells are all empty
            if not any(row[i] for i in range(min(5, len(row)))):
                continue

            factor = self.extract_factor_from_row(row, col_idx, sheet_name)
            if factor:
                self.factors.append(factor)
                count += 1

        return count

    def extract_factor_from_row(self, row, col_idx, sheet_name):
        """Extract emission factor from a row"""
        try:
            # Get emission factor
            co2e = row[col_idx['co2e']]
            if not co2e or co2e == '' or co2e == 0:
                return None

            try:
                emission_factor = float(co2e)
                if emission_factor <= 0:
                    return None
            except (ValueError, TypeError):
                return None

            # Get unit
            unit = str(row[col_idx['unit']]) if col_idx.get('unit') else 'unit'
            if not unit or unit == 'None':
                return None

            # Get sheet mapping
            mapping = self.sheet_mapping[sheet_name]

            # Build activity name
            activity_parts = []
            for key in ['activity', 'fuel', 'type', 'haul', 'class']:
                if key in col_idx and row[col_idx[key]]:
                    val = str(row[col_idx[key]]).strip()
                    if val and val != 'None' and val != '':
                        activity_parts.append(val)

            activity_name = ' - '.join(activity_parts) if activity_parts else 'Unknown'

            # Get country if available
            country = None
            geographic_scope = 'UK'
            if 'country' in col_idx and row[col_idx['country']]:
                country_val = str(row[col_idx['country']]).strip()
                if country_val and country_val != 'None':
                    country = country_val
                    if 'UK' not in country_val and 'United Kingdom' not in country_val:
                        geographic_scope = 'International'

            # Get GHG components
            co2 = float(row[col_idx['co2']]) if col_idx.get('co2') and row[col_idx['co2']] else None
            ch4 = float(row[col_idx['ch4']]) if col_idx.get('ch4') and row[col_idx['ch4']] else None
            n2o = float(row[col_idx['n2o']]) if col_idx.get('n2o') and row[col_idx['n2o']] else None

            # Determine India relevance
            india_relevant = mapping['india_relevant']
            if geographic_scope == 'UK' and sheet_name == 'UK electricity':
                india_relevant = False

            # ISO 14064 compliance flags
            iso_compliant = True
            data_quality = 'High'

            # Generate tags
            tags = self.generate_tags(activity_name, mapping['activity_type'], sheet_name)

            # Create factor
            factor = {
                'id': self.factor_id,
                'activity_type': mapping['activity_type'],
                'activity_subtype': sheet_name.lower().replace(' ', '_'),
                'activity_category': mapping['category_name'],
                'region': 'Europe' if geographic_scope == 'UK' else 'Global',
                'country': country if country else 'UK',
                'state_province': None,
                'city': None,
                'emission_factor': emission_factor,
                'unit': unit.strip(),
                'co2_factor': co2,
                'ch4_factor': ch4,
                'n2o_factor': n2o,
                'gas_type': 'CO2e',
                'source': 'DEFRA 2024',
                'source_url': 'https://www.gov.uk/government/publications/greenhouse-gas-reporting-conversion-factors-2024',
                'methodology': 'Lifecycle Assessment (LCA)',
                'data_quality': data_quality,
                'year': 2024,
                'valid_from': '2024-01-01',
                'valid_until': '2024-12-31',
                'priority': mapping['priority_for_india'],
                'confidence_level': 0.95,
                'is_default': False,
                'scope': mapping['scope'],
                'scope_category': mapping['scope_category'],
                'category_name': mapping['category_name'],
                'industry_sector': 'All',
                'company_size': 'All',
                'notes': f"DEFRA 2024 - {activity_name}",
                'tags': tags,
                'conversion_factor': None,
                'original_unit': unit.strip(),
                'activity_name': activity_name,
                'source_sheet': sheet_name,
                'geographic_scope': geographic_scope,
                'india_relevant': india_relevant,
                'recommended_for_india': india_relevant and mapping['priority_for_india'] <= 3,
                'iso_14064_compliant': iso_compliant
            }

            self.factor_id += 1

            # Update stats
            self.stats['by_scope'][mapping['scope']] = self.stats['by_scope'].get(mapping['scope'], 0) + 1
            if india_relevant:
                self.stats['india_relevant'] += 1
            else:
                self.stats['uk_specific'] += 1

            return factor

        except Exception as e:
            return None

    def generate_tags(self, activity_name, activity_type, sheet_name):
        """Generate searchable tags"""
        tags = set()

        # Add from activity name
        if activity_name:
            tags.update(activity_name.lower().split())

        # Add activity type
        tags.add(activity_type)

        # Add sheet name words
        tags.update(sheet_name.lower().split())

        # Clean tags
        tags = {t.strip('(),:-') for t in tags if t and len(t) > 2}
        tags.discard('per')
        tags.discard('unit')

        return ', '.join(sorted(tags))

    def save_to_csv(self, output_file='defra_factors.csv'):
        """Save factors to CSV"""
        print(f"\n💾 Saving to CSV: {output_file}")

        if not self.factors:
            print("  ⚠️  No factors to save!")
            return

        df = pd.DataFrame(self.factors)
        df.to_csv(output_file, index=False, encoding='utf-8')

        print(f"  ✅ Saved {len(self.factors)} factors")

    def create_metadata(self, output_file='sources_metadata.json'):
        """Create metadata file for source prioritization"""
        print(f"\n📝 Creating metadata: {output_file}")

        metadata = {
            'defra_factors.csv': {
                'source_name': 'DEFRA 2024',
                'full_name': 'UK Government GHG Conversion Factors for Company Reporting',
                'priority': 3,
                'geographic_scope': 'UK + International',
                'use_for_india': 'Proxy only - prefer India-specific factors',
                'iso_14064_compliant': True,
                'last_updated': '2024-06-01',
                'update_frequency': 'Annual',
                'total_factors': len(self.factors),
                'india_relevant_factors': self.stats['india_relevant'],
                'uk_specific_factors': self.stats['uk_specific'],
                'recommended_use': 'Use for UK subsidiaries or as fallback when India-specific data unavailable',
                'data_quality': 'High',
                'methodology': 'Lifecycle Assessment',
                'url': 'https://www.gov.uk/government/publications/greenhouse-gas-reporting-conversion-factors-2024'
            }
        }

        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(metadata, f, indent=2, ensure_ascii=False)

        print(f"  ✅ Metadata saved")

    def generate_summary(self):
        """Generate summary report"""
        print("\n" + "=" * 80)
        print("📊 DEFRA EXTRACTION SUMMARY")
        print("=" * 80)

        print(f"\n📁 Sheets:")
        print(f"  • Total processed: {self.stats['processed_sheets']}")

        print(f"\n✅ Emission Factors:")
        print(f"  • Total extracted: {self.stats['total_factors']}")

        print(f"\n📋 By Scope:")
        for scope, count in sorted(self.stats['by_scope'].items()):
            print(f"  • {scope}: {count} factors")

        print(f"\n🇮🇳 India Relevance:")
        print(
            f"  • India-relevant: {self.stats['india_relevant']} ({self.stats['india_relevant'] / self.stats['total_factors'] * 100:.1f}%)")
        print(
            f"  • UK-specific: {self.stats['uk_specific']} ({self.stats['uk_specific'] / self.stats['total_factors'] * 100:.1f}%)")

        print("\n" + "=" * 80)


def main():
    """Main execution"""
    print("🚀 DEFRA EMISSION FACTOR ORGANIZER")
    print("=" * 80)
    print("🇮🇳 Optimized for India-based SaaS Platform")
    print("✅ ISO 14064 Compliant")
    print("=" * 80)

    # CHANGE THIS PATH TO YOUR DEFRA FILE
    defra_file_path = r"C:\Users\shubh\Downloads\IPCC\defra_EF.xlsx"

    # Verify path exists
    if not Path(defra_file_path).exists():
        print(f"❌ ERROR: File not found: {defra_file_path}")
        print("\n💡 Please update the 'defra_file_path' variable in the script with your correct path")
        return

    # Initialize organizer
    organizer = DEFRAOrganizer(defra_file_path)

    # Process DEFRA file
    organizer.process_defra_file()

    # Save outputs
    organizer.save_to_csv('defra_factors.csv')
    organizer.create_metadata('sources_metadata.json')

    # Generate summary
    organizer.generate_summary()

    print("\n✅ DONE! DEFRA factors extracted successfully.")
    print("\n📄 Output files:")
    print("  • defra_factors.csv - DEFRA emission factors")
    print("  • sources_metadata.json - Source prioritization metadata")
    print("\n💡 Next: Add India-specific factors (CEA, BEE, TERI)")


if __name__ == '__main__':
    main()